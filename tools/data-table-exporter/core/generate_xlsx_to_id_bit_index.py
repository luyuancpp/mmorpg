#!/usr/bin/env python
# coding=utf-8

import json
import logging
import concurrent.futures
import openpyxl
import multiprocessing
from typing import List, Optional, Dict
from pathlib import Path

import generate_common
from core import paths
from jinja2 import Environment, FileSystemLoader

# Setup Logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelToCppConverter:
    def __init__(self, excel_file: Path):
        self.excel_file = excel_file
        self.workbook = openpyxl.load_workbook(excel_file)
        self.sheet = self.workbook.sheetnames[0]
        self.worksheet = self.workbook[self.sheet]
        self.bit_index_col = self._find_bit_index_column()
        self.mapping_file = paths.GENERATOR_STORAGE_TABLE_INDEX_DIR / f"{self.sheet.lower()}_mapping.json"

        # Initialize Jinja2 environment
        self.template_env = Environment(
            loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'),
            autoescape=True
        )

    def _find_bit_index_column(self) -> Optional[int]:
        headers = [cell.value for cell in self.worksheet[1]]
        for col_idx in range(len(headers)):
            cell_value = self.worksheet.cell(row=generate_common.XLSX_TABLE_BIT_BEGIN_INDEX, column=col_idx + 1).value
            if cell_value is not None and cell_value.strip().lower() == 'bit_index':
                return col_idx
        return None

    def _load_existing_mapping(self) -> Dict[int, int]:
        if self.mapping_file.exists():
            try:
                with self.mapping_file.open('r', encoding='utf-8') as file:
                    data = json.load(file)
                    return {int(k): v for k, v in data.items()}
            except json.JSONDecodeError:
                logger.error("Error: JSON file is not valid.")
        return {}

    def _save_mapping(self, mapping: Dict[int, int]) -> None:
        with self.mapping_file.open('w', encoding='utf-8') as file:
            json.dump(mapping, file, indent=4)

    def _find_unused_indexes(self, id_to_index: Dict[int, int]) -> List[int]:
        used_indexes = set(id_to_index.values())
        all_indexes = set(range(len(id_to_index)))
        return sorted(all_indexes - used_indexes)

    def _find_max_bit_index(self) -> int:
        max_bit_index = -1
        if self.bit_index_col is not None:
            for row in self.worksheet.iter_rows(min_row=20, values_only=True):
                bit_index = row[self.bit_index_col]
                if isinstance(bit_index, int) and bit_index > max_bit_index:
                    max_bit_index = bit_index
        return max_bit_index

    def should_process(self) -> bool:
        return self.bit_index_col is not None

    def generate_cpp_constants(self) -> str:
        id_to_index = self._load_existing_mapping()
        unused_indexes = self._find_unused_indexes(id_to_index)
        current_index = max(id_to_index.values(), default=-1) + 1

        for row in self.worksheet.iter_rows(min_row=20, values_only=True):
            id_value = row[0]
            if id_value is None:
                continue
            if id_value not in id_to_index:
                index = unused_indexes.pop(0) if unused_indexes else current_index
                current_index += 1 if not unused_indexes else 0
                id_to_index[id_value] = index

        template = self.template_env.get_template("cpp_config_id_bit_template.h.j2")
        cpp_constants = template.render(
            sheet=self.sheet,
            id_to_index=id_to_index,
            max_bit_index=self._find_max_bit_index()
        )

        self._save_mapping(id_to_index)
        return cpp_constants

    def generate_go_constants(self) -> str:
        id_to_index = self._load_existing_mapping()
        unused_indexes = self._find_unused_indexes(id_to_index)
        current_index = max(id_to_index.values(), default=-1) + 1

        for row in self.worksheet.iter_rows(min_row=20, values_only=True):
            id_value = row[0]
            if id_value is None:
                continue
            if id_value not in id_to_index:
                index = unused_indexes.pop(0) if unused_indexes else current_index
                current_index += 1 if not unused_indexes else 0
                id_to_index[id_value] = index

        template = self.template_env.get_template("go_config_id_bit_template.go.j2")
        go_constants = template.render(
            sheet=self.sheet,
            id_to_index=id_to_index,
            max_bit_index=self._find_max_bit_index()
        )

        self._save_mapping(id_to_index)
        return go_constants

    def save_go_constants_to_file(self, go_constants: str) -> None:
        output_file = paths.SRC_GO_ID_BIT / f"{self.sheet.lower()}_table_id_bit_index.go"
        output_file.write_text(go_constants, encoding='utf-8')

    def save_cpp_constants_to_file(self, cpp_constants: str) -> None:
        output_file = paths.SRC_CPP_ID_BIT_INDEX / f"{self.sheet.lower()}_table_id_bit_index.h"
        output_file.write_text(cpp_constants, encoding='utf-8')


def get_xlsx_files(directory: Path) -> List[Path]:
    return [file for file in directory.glob("*.xlsx") if file.is_file()]


def process_file(excel_file: Path) -> None:
    converter = ExcelToCppConverter(excel_file)
    if converter.should_process():
        cpp_constants = converter.generate_cpp_constants()
        converter.save_cpp_constants_to_file(cpp_constants)

        go_constants = converter.generate_go_constants()
        converter.save_go_constants_to_file(go_constants)
    else:
        logger.debug(f"Skipping file {excel_file} due to missing 'bit_index'.")


def main() -> None:
    paths.SRC_CPP_ID_BIT_INDEX.mkdir(parents=True, exist_ok=True)
    paths.SRC_GO_ID_BIT.mkdir(parents=True, exist_ok=True)
    paths.GENERATOR_STORAGE_TABLE_INDEX_DIR.mkdir(parents=True, exist_ok=True)

    try:
        xlsx_files = get_xlsx_files(paths.DATA_TABLES_DIR)
        num_threads = min(multiprocessing.cpu_count(), len(xlsx_files))

        with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
            futures = [executor.submit(process_file, file_path) for file_path in xlsx_files]
            for future in concurrent.futures.as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    logger.error(f"任务执行失败: {str(e)}")
    except Exception as e:
        logger.error(f"主程序执行失败: {str(e)}")


if __name__ == "__main__":
    main()
