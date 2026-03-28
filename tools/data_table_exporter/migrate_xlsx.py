"""Migrate all .xlsx files from old 19-row header to new 6-row compact header.

Old format (rows 1–19):
    1: column_name   2: data_type   3: map_type (set|map_key|map_value)
    4: owner   5: multi_key   6: table_key / bit_index
    7: expression_type   8: expression_params   9: foreign_key
    10: group_foreign_key   11–18: (empty)   19: comment
    20+: data

New format (rows 1–6):
    1: column_name   2: data_type
    3: struct (repeated|set|map_key|map_value|message:Name)
    4: owner
    5: options (space-separated: bit_index key multi fk:T gfk:T expr:type expr_params:a,b)
    6: comment
    7+: data

Usage:
    python migrate_xlsx.py [--dry-run]
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
OLD_DATA_BEGIN = 20
NEW_DATA_BEGIN = 7


def cell_str(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def compute_structs(names: list[str], map_types: list[str]) -> list[str]:
    """Compute the new struct row from old map_type markers + pattern detection."""
    n = len(names)
    structs = [""] * n

    # Step 1: copy explicit markers
    for i in range(n):
        mt = map_types[i].lower()
        if mt in ("set", "map_key", "map_value"):
            structs[i] = mt

    # Step 2: propagate "set" to ALL columns with the same name
    set_names = {names[i] for i in range(n) if structs[i] == "set" and names[i]}
    for i in range(n):
        if names[i] in set_names and not structs[i]:
            structs[i] = "set"

    # Step 3: propagate map_key/map_value to all matching name-pairs
    map_pairs: set[tuple[str, str]] = set()
    for i in range(n - 1):
        if structs[i] == "map_key" and structs[i + 1] == "map_value":
            map_pairs.add((names[i], names[i + 1]))
    for i in range(n - 1):
        if not structs[i] and not structs[i + 1]:
            if (names[i], names[i + 1]) in map_pairs:
                structs[i] = "map_key"
                structs[i + 1] = "map_value"

    # Step 4: detect repeated (consecutive same-name) and message groups
    # First, find all columns still untagged
    untagged = {i for i in range(n) if not structs[i] and names[i]}

    # Find consecutive same-name runs among untagged columns
    i = 0
    while i < n:
        if i not in untagged:
            i += 1
            continue
        j = i + 1
        while j < n and j in untagged and names[j] == names[i]:
            j += 1
        if j - i > 1:
            for k in range(i, j):
                structs[k] = "repeated"
                untagged.discard(k)
        i = j

    # Find non-consecutive same-name patterns (message groups)
    # Collect untagged names that appear more than once
    name_positions: dict[str, list[int]] = {}
    for i in sorted(untagged):
        name_positions.setdefault(names[i], []).append(i)

    repeated_names = {nm for nm, positions in name_positions.items() if len(positions) > 1}
    if repeated_names:
        # Find contiguous ranges that contain repeated names
        groups = _find_message_groups(names, structs, repeated_names, untagged)
        for group_name, indices in groups.items():
            for idx in indices:
                structs[idx] = f"message:{group_name}"

    return structs


def _find_message_groups(
    names: list[str],
    structs: list[str],
    repeated_names: set[str],
    untagged: set[int],
) -> dict[str, list[int]]:
    """Find message groups from non-consecutive same-name columns.

    Uses range-merging: each repeated name defines a [min, max] range of
    its positions.  Overlapping ranges are merged into one message group.
    """
    n = len(names)

    # Step 1: find [min, max] range for each repeated name
    name_ranges: dict[str, tuple[int, int]] = {}
    for target_name in repeated_names:
        positions = [i for i in range(n) if names[i] == target_name and i in untagged]
        if len(positions) >= 2:
            name_ranges[target_name] = (min(positions), max(positions))

    if not name_ranges:
        return {}

    # Step 2: merge overlapping ranges (strict overlap, not adjacent)
    sorted_ranges = sorted(name_ranges.values())
    merged: list[tuple[int, int]] = []
    cur_start, cur_end = sorted_ranges[0]
    for start, end in sorted_ranges[1:]:
        if start <= cur_end:          # overlapping
            cur_end = max(cur_end, end)
        else:
            merged.append((cur_start, cur_end))
            cur_start, cur_end = start, end
    merged.append((cur_start, cur_end))

    # Step 3: for each merged range, collect untagged columns → one group
    groups: dict[str, list[int]] = {}
    for rng_start, rng_end in merged:
        indices = [i for i in range(rng_start, rng_end + 1) if i in untagged]
        col_names = [names[i] for i in indices]
        prefix = _common_prefix_multi(col_names)
        if not prefix:
            prefix = col_names[0]
        # Avoid duplicate group names
        final_name = prefix
        counter = 2
        while final_name in groups:
            final_name = f"{prefix}{counter}"
            counter += 1
        groups[final_name] = indices

    return groups


def _common_prefix_multi(names: list[str]) -> str:
    """Find common underscore-delimited prefix across multiple names."""
    if not names:
        return ""
    parts_list = [n.split("_") for n in names]
    min_len = min(len(p) for p in parts_list)
    common = []
    for i in range(min_len):
        vals = {p[i] for p in parts_list}
        if len(vals) == 1:
            common.append(vals.pop())
        else:
            break
    return "_".join(common)


def compute_options(
    multi_keys: list[str],
    table_keys: list[str],
    expr_types: list[str],
    expr_params: list[str],
    fk_vals: list[str],
    gfk_vals: list[str],
) -> list[str]:
    """Compute the new options row from old scattered metadata rows."""
    n = len(multi_keys)
    options = [""] * n

    for i in range(n):
        tokens: list[str] = []

        if multi_keys[i].lower() == "multi":
            tokens.append("multi")
        if table_keys[i].lower() == "table_key":
            tokens.append("key")
        if table_keys[i].lower() == "bit_index":
            tokens.append("bit_index")
        if expr_types[i]:
            tokens.append(f"expr:{expr_types[i]}")
        if expr_params[i]:
            tokens.append(f"expr_params:{expr_params[i]}")
        if fk_vals[i]:
            tokens.append(fk_vals[i])        # already "fk:Table" format
        if gfk_vals[i]:
            tokens.append(gfk_vals[i])      # already "gfk:Table" format

        options[i] = " ".join(tokens)

    return options


def migrate_file(path: Path, dry_run: bool = False) -> None:
    """Migrate a single .xlsx file from old to new header format."""
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    max_col = ws.max_column or 0

    if max_col == 0:
        print(f"  SKIP (empty): {path.name}")
        return

    # Read old header rows
    names = [cell_str(ws, 1, c + 1) for c in range(max_col)]
    types = [cell_str(ws, 2, c + 1) for c in range(max_col)]
    map_types = [cell_str(ws, 3, c + 1) for c in range(max_col)]
    owners = [cell_str(ws, 4, c + 1) for c in range(max_col)]
    multi_keys = [cell_str(ws, 5, c + 1) for c in range(max_col)]
    table_keys = [cell_str(ws, 6, c + 1) for c in range(max_col)]
    expr_types = [cell_str(ws, 7, c + 1) for c in range(max_col)]
    expr_params = [cell_str(ws, 8, c + 1) for c in range(max_col)]
    fk_vals = [cell_str(ws, 9, c + 1) for c in range(max_col)]
    gfk_vals = [cell_str(ws, 10, c + 1) for c in range(max_col)]
    comments = [cell_str(ws, 19, c + 1) for c in range(max_col)]

    # Read data rows (preserve raw cell values)
    data_rows: list[list] = []
    for row in ws.iter_rows(min_row=OLD_DATA_BEGIN, max_col=max_col, values_only=True):
        data_rows.append(list(row))

    # Remove trailing empty data rows
    while data_rows and all(v is None for v in data_rows[-1]):
        data_rows.pop()

    # Compute new struct and options
    structs = compute_structs(names, map_types)
    options = compute_options(multi_keys, table_keys, expr_types, expr_params, fk_vals, gfk_vals)

    if dry_run:
        print(f"\n  {path.name} ({len(data_rows)} data rows, {max_col} cols)")
        for c in range(max_col):
            if not names[c]:
                continue
            parts = [f"name={names[c]}", f"type={types[c]}"]
            if structs[c]:
                parts.append(f"struct={structs[c]}")
            if options[c]:
                parts.append(f"options={options[c]}")
            if comments[c]:
                parts.append(f"comment={comments[c][:30]}")
            print(f"    col {c}: {', '.join(parts)}")
        return

    # Write new workbook
    sheet_title = ws.title
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_title

    for c in range(max_col):
        new_ws.cell(row=1, column=c + 1, value=names[c] if names[c] else None)
        new_ws.cell(row=2, column=c + 1, value=types[c] if types[c] else None)
        new_ws.cell(row=3, column=c + 1, value=structs[c] if structs[c] else None)
        new_ws.cell(row=4, column=c + 1, value=owners[c] if owners[c] else None)
        new_ws.cell(row=5, column=c + 1, value=options[c] if options[c] else None)
        new_ws.cell(row=6, column=c + 1, value=comments[c] if comments[c] else None)

    for r, row_data in enumerate(data_rows):
        for c, val in enumerate(row_data):
            if val is not None:
                new_ws.cell(row=NEW_DATA_BEGIN + r, column=c + 1, value=val)

    new_wb.save(path)
    print(f"  OK: {path.name} ({len(data_rows)} data rows)")


def main() -> None:
    dry_run = "--dry-run" in sys.argv

    xlsx_files = sorted(DATA_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No .xlsx files found in {DATA_DIR}")
        return

    mode = "DRY RUN" if dry_run else "MIGRATING"
    print(f"{mode}: {len(xlsx_files)} files in {DATA_DIR}")

    for f in xlsx_files:
        if f.name.startswith("~$"):
            continue
        try:
            migrate_file(f, dry_run=dry_run)
        except Exception as exc:
            print(f"  FAIL: {f.name}: {exc}")


if __name__ == "__main__":
    main()
