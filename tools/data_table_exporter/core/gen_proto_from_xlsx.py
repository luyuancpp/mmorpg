#!/usr/bin/env python
# coding=utf-8

import openpyxl
import logging
import concurrent.futures
import multiprocessing
from typing import Dict, List, Optional
from pathlib import Path
import generate_common
from constants import DATA_TABLES_DIR, PROTO_DIR

# Setup Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def get_workbook_data(workbook: openpyxl.Workbook) -> Dict[str, Dict]:
    """Extract data from the first sheet in the workbook."""
    data = {}
    if workbook.sheetnames:
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        if validate_sheet(sheet):
            column_names = generate_common.get_column_names(sheet)
            if column_names:
                sheet_data = generate_common.get_sheet_data(sheet, column_names)
                data[sheet_name] = sheet_data
    else:
        logger.error("No sheets found in the workbook.")
    return data

def validate_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    """Ensure the sheet's first column header is 'id'."""
    if sheet.cell(row=1, column=1).value != "id":
        logger.error(f"Sheet '{sheet.title}' first column must be 'id'")
        return False
    return True

def generate_proto_file(data: Dict, sheet_name: str) -> Optional[str]:
    """Generate .proto file content based on sheet data."""
    try:
        proto_content = create_proto_header()
        column_names = data[generate_common.SHEET_COLUM_NAME_INDEX]

        proto_content += generate_group_messages(sheet_name, data, column_names)
        proto_content += generate_row_message(sheet_name, data, column_names)
        proto_content += generate_table_message(sheet_name)

        return proto_content
    except Exception as e:
        logger.error(f"Error generating proto content for '{sheet_name}': {e}")
        return None

def create_proto_header() -> str:
    """Create the initial header for the .proto file."""
    return (
        'syntax = "proto3";\n\n'
        'option go_package = "pb/game";\n\n'
    )

def generate_group_messages(sheet_name: str, data: Dict, column_names: List[str]) -> str:
    """Generate messages for grouped data."""
    proto_content = ''
    group_data = data[generate_common.SHEET_GROUP_ARRAY_DATA_INDEX]
    for k, v in group_data.items():
        obj_name = sheet_name + generate_common.set_to_string(
            generate_common.find_common_words(column_names[v[0]], column_names[v[1]], '_')
        )
        proto_content += f'message {obj_name} {{\n'
        for i, column in enumerate(v):
            name = column_names[column]
            proto_content += f'\t{data[0][name]} {name} = {i + 1};\n'
        proto_content += '}\n\n'
    return proto_content

def generate_row_message(sheet_name: str, data: Dict, column_names: List[str]) -> str:
    """Generate the row message for the .proto file."""
    proto_content = f'message {sheet_name}Table {{\n'
    field_index = 1

    for key, _ in data[0].items():
        if not is_excluded_owner(data[generate_common.OWNER_INDEX], key):
            field_content = format_field(sheet_name, data, key, column_names, field_index)
            if field_content:
                proto_content += field_content
                field_index += 1

    proto_content += '}\n\n'
    return proto_content

def is_excluded_owner(owner_data: Dict, key: str) -> bool:
    """Check if the key is excluded based on owner data."""
    return owner_data.get(key, '').strip() in ('client', 'design')

def format_field(sheet_name: str, data: Dict, key: str, column_names: List[str], field_index: int) -> str:
    """Format a field for the .proto file based on its type."""
    if key in data[generate_common.MAP_TYPE_INDEX]:
        return format_map_field(data, key, column_names, field_index)
    elif key in data[generate_common.SHEET_ARRAY_DATA_INDEX]:
        return f'\trepeated {data[0][key]} {key} = {field_index};\n'
    elif generate_common.is_key_in_group_array(data[generate_common.SHEET_GROUP_ARRAY_DATA_INDEX], key, column_names):
        return format_group_array_field(sheet_name, data, key, column_names, field_index)
    else:
        return f'\t{data[0][key]} {key} = {field_index};\n'

def format_map_field(data: Dict, key: str, column_names: List[str], field_index: int) -> str:
    """Format a map field for the .proto file."""
    map_type = data[generate_common.MAP_TYPE_INDEX][key]
    if map_type == generate_common.SET_CELL:
        return f'\tmap <{data[0][key]}, bool> {key} = {field_index};\n'
    elif map_type == generate_common.MAP_KEY_CELL:
        value_type = data[generate_common.FILE_TYPE_INDEX]
        if key in data[generate_common.SHEET_GROUP_ARRAY_DATA_INDEX]:
            value = data[generate_common.SHEET_GROUP_ARRAY_DATA_INDEX][key]
            key_name = column_names[value[0]]
            value_name = column_names[value[1]]
            obj_name = generate_common.column_name_to_obj_name(key_name, '_')
            return f'\tmap <{value_type[key_name]}, {value_type[value_name]}> {obj_name} = {field_index};\n'
    return ''

def format_group_array_field(sheet_name: str, data: Dict, key: str, column_names: List[str], field_index: int) -> str:
    """Format a group array field for the .proto file."""
    if key not in data[generate_common.SHEET_GROUP_ARRAY_DATA_INDEX]:
        return ''
    value = data[generate_common.SHEET_GROUP_ARRAY_DATA_INDEX][key]
    key_name = column_names[value[0]]
    obj_name = generate_common.column_name_to_obj_name(key_name, '_')
    return f'\trepeated {sheet_name}{obj_name} {obj_name} = {field_index};\n'

def generate_table_message(sheet_name: str) -> str:
    """Generate the table message for the .proto file."""
    return (
        f'message {sheet_name}TabledData {{\n'
        f'\trepeated {sheet_name}Table data = 1;\n'
        '}\n'
    )

def process_file(file_path: str) -> None:
    """Process an individual Excel file to generate .proto files."""
    try:
        workbook = openpyxl.load_workbook(file_path)
        workbook_data = get_workbook_data(workbook)

        for sheet_name, data in workbook_data.items():
            proto_content = generate_proto_file(data, sheet_name)
            if proto_content:
                save_proto_file(proto_content, sheet_name.lower())
    except Exception as e:
        logger.error(f"Error processing file '{file_path}': {e}")

def save_proto_file(content: str, sheet_name: str) -> None:
    """Save the generated .proto content to a file."""
    proto_file_path = PROTO_DIR / f'{sheet_name}_table.proto'
    try:
        with proto_file_path.open('w', encoding='utf-8') as proto_file:
            proto_file.write(content)
            logger.info(f"Generated .proto file: {proto_file_path}")
    except Exception as e:
        logger.error(f"Error saving proto file '{proto_file_path}': {e}")

def get_xlsx_files(directory: Path) -> List[str]:
    """List all .xlsx files in the specified directory."""
    return [str(f) for f in directory.glob('*.xlsx') if f.is_file()]

def main() -> None:
    """Main function to process all Excel files and generate .proto files."""
    try:
        PROTO_DIR.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        logger.error(f"Failed to create proto directory: {e}")
        return

    try:
        xlsx_files = get_xlsx_files(DATA_TABLES_DIR)
    except Exception as e:
        logger.error(f"Failed to list .xlsx files: {e}")
        return

    num_threads = min(multiprocessing.cpu_count(), len(xlsx_files))
    with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = [executor.submit(process_file, file_path) for file_path in xlsx_files]
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logger.error(f"An error occurred during processing: {e}")

if __name__ == "__main__":
    main()
