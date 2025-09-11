#!/usr/bin/env python
# coding=utf-8

import openpyxl
import generate_common  # Ensure generate_common provides BEGIN_ROW_IDX and mywrite functions
import logging
from pathlib import Path
from core.paths import DATA_TABLES_DIR

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

cpp_dir = Path("generated/cpp/")

def generate_id_enum(sheet):
    """Generate a C++ enum definition based on the provided sheet."""
    n_rows = sheet.max_row
    name = sheet.title
    file_str = "#pragma once\n"
    file_str += f"enum e_{name}_table_id : uint32_t\n{{\n"

    for idx in range(generate_common.BEGIN_ROW_IDX, n_rows + 1):
        cell_value = sheet.cell(row=idx, column=1).value
        if cell_value is not None:
            file_str += f"    {name}_table_id{int(cell_value)},\n"

    file_str += "};\n"
    return file_str

def generate_id_cpp(workbook):
    """Generate C++ enum files for the first sheet in the workbook."""
    workbook_data = {}
    first_sheet_name = workbook.sheetnames[0]  # Only get the first sheet
    if first_sheet_name in generate_common.GEN_FILE_LIST:
        sheet = workbook[first_sheet_name]
        file_data = generate_id_enum(sheet)
        workbook_data[first_sheet_name] = file_data
    return workbook_data

def main():
    """Main function to process all Excel files and generate C++ enum files."""
    if not cpp_dir.exists():
        cpp_dir.mkdir(parents=True, exist_ok=True)

    for filepath in DATA_TABLES_DIR.glob("*.xlsx"):
        if filepath.is_file():
            try:
                workbook = openpyxl.load_workbook(filepath)
                workbook_data = generate_id_cpp(workbook)
                for sheet_name, data in workbook_data.items():
                    output_file_path = cpp_dir / f"{sheet_name.lower()}_table_id.h"
                    generate_common.mywrite(data, str(output_file_path))
                    logger.info(f"Generated C++ enum file: {output_file_path}")
            except Exception as e:
                logger.error(f"Failed to process file {filepath}: {e}")

if __name__ == "__main__":
    main()
