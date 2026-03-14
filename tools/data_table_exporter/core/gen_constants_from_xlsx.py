#!/usr/bin/env python
# coding=utf-8

import logging
import openpyxl
import sys
import io
from pathlib import Path
from typing import Optional
from jinja2 import Environment, FileSystemLoader

# Ensure console output supports UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

import utils
import generate_common
from core import paths

# Set up logging configuration
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ExcelConstantsGenerator:
    def __init__(self, excel_file: Path):
        self.excel_file = excel_file
        self.is_global_file = 'globalvariable' in self.excel_file.name.lower()
        try:
            self.workbook = openpyxl.load_workbook(self.excel_file)
            self.sheet = self.workbook.sheetnames[0]
            self.worksheet = self.workbook[self.sheet]
            self.constants_name_index = self._find_constants_name_index()
        except Exception as e:
            logger.error(f"Failed to load Excel file: {self.excel_file}, error: {e}")
            raise

    def _find_constants_name_index(self) -> Optional[int]:
        try:
            if not self.worksheet or self.worksheet.max_row < 1:
                logger.warning(f"Worksheet '{self.sheet}' is empty or has no data")
                return None

            first_row = self.worksheet[1]
            headers = [str(cell.value).strip() if cell.value else '' for cell in first_row]
            return headers.index('constants_name') if 'constants_name' in headers else None
        except Exception as e:
            logger.error(f"Error finding 'constants_name' column: {e}")
            return None

    def should_process(self) -> bool:
        return self.constants_name_index is not None

    def generate_cpp_constants(self) -> str:
        constants_list = []
        for row in self.worksheet.iter_rows(min_row=20, values_only=True):
            id_value = row[0]
            if id_value is None:
                continue
            constant_name = self._generate_constant_name(row, id_value)
            constants_list.append({'name': constant_name, 'value': id_value})

        env = Environment(loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'))
        template = env.get_template("constants.h.j2")
        return template.render(constants=constants_list)

    def _generate_constant_name(self, row: tuple, id_value: int) -> str:
        if self.constants_name_index is not None and row[self.constants_name_index]:
            return f'k{self.sheet}_{row[self.constants_name_index]}'
        return f'k{self.sheet}_{id_value}'

    def close(self):
        if hasattr(self, 'workbook'):
            self.workbook.close()

    def generate_and_save_cpp_constants(self):
        if not self.should_process():
            logger.info(f"No 'constants_name' column found, skipping file: {self.excel_file}")
            return

        env = Environment(loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'))
        template = env.get_template("constants.h.j2")

        if self.is_global_file:
            constants_map = {}
            single_list = []

            for row in self.worksheet.iter_rows(min_row=20, values_only=True):
                id_value = row[0]
                const_raw = row[self.constants_name_index]
                if id_value is None:
                    continue

                const_name = str(const_raw) if const_raw else str(id_value)

                if '_' in const_name:
                    parts = const_name.split('_')
                    if len(parts) >= 2:
                        key = (parts[0], parts[1])
                        cpp_name = f'k{self.sheet}_{const_name}'
                        constants_map.setdefault(key, []).append({'name': cpp_name, 'value': id_value})
                    else:
                        logger.warning(f"Invalid underscore format in: {const_name}")
                else:
                    cpp_name = f'k{self.sheet}_{const_name}'
                    single_list.append({'name': cpp_name, 'value': id_value})

            for key, const_list in constants_map.items():
                first, second = key
                cpp_code = template.render(constants=const_list)
                filename = f"global_{first}_{second}_table_id_constants.h".lower()
                output_path = paths.SRC_CPP_CONSTANTS / filename
                output_path.write_text(cpp_code, encoding='utf-8')
                logger.info(f"Generated file: {output_path}")

            if single_list:
                cpp_code = template.render(constants=single_list)
                filename = f"{self.sheet.lower()}_table_id_constants.h"
                output_path = paths.SRC_CPP_CONSTANTS / filename
                output_path.write_text(cpp_code, encoding='utf-8')
                logger.info(f"Generated fallback file: {output_path}")

        else:
            constants_list = []
            for row in self.worksheet.iter_rows(min_row=20, values_only=True):
                id_value = row[0]
                if id_value is None:
                    continue
                const_raw = row[self.constants_name_index]
                if not const_raw:
                    continue
                cpp_name = f'k{self.sheet}_{const_raw}'
                constants_list.append({'name': cpp_name, 'value': id_value})

            cpp_code = template.render(constants=constants_list)
            output_path = paths.SRC_CPP_CONSTANTS / f"{self.sheet.lower()}_table_id_constants.h"
            output_path.write_text(cpp_code, encoding='utf-8')
            logger.info(f"Generated file: {output_path}")

    def generate_and_save_go_constants(self):
        if not self.should_process():
            logger.info(f"[GO] No 'constants_name' column found, skipping: {self.excel_file}")
            return

        env = Environment(loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'))
        template = env.get_template("constants.go.j2")

        if self.is_global_file:
            constants_map = {}
            single_list = []

            for row in self.worksheet.iter_rows(min_row=20, values_only=True):
                id_value = row[0]
                const_raw = row[self.constants_name_index]
                if id_value is None:
                    continue

                const_name = str(const_raw) if const_raw else str(id_value)

                if '_' in const_name:
                    parts = const_name.split('_')
                    if len(parts) >= 2:
                        key = (parts[0], parts[1])
                        go_const_name = f'K{self.sheet}_{const_name}'
                        constants_map.setdefault(key, []).append({'name': go_const_name, 'value': id_value})
                    else:
                        logger.warning(f"[GO] Invalid underscore format in: {const_name}")
                else:
                    go_const_name = f'K{self.sheet}_{const_name}'
                    single_list.append({'name': go_const_name, 'value': id_value})

            for key, const_list in constants_map.items():
                first, second = key
                go_code = template.render(constants=const_list)
                filename = f"global_{first}_{second}_table_id_constants.go".lower()
                output_path = paths.SRC_GO_CONSTANTS / filename
                output_path.write_text(go_code, encoding='utf-8')
                logger.info(f"[GO] Generated file: {output_path}")

            if single_list:
                go_code = template.render(constants=single_list)
                filename = f"{self.sheet.lower()}_table_id_constants.go"
                output_path = paths.SRC_GO_CONSTANTS / filename
                output_path.write_text(go_code, encoding='utf-8')
                logger.info(f"[GO] Generated fallback file: {output_path}")

        else:
            constants_list = []
            for row in self.worksheet.iter_rows(min_row=20, values_only=True):
                id_value = row[0]
                if id_value is None:
                    continue
                const_raw = row[self.constants_name_index]
                if not const_raw:
                    continue
                go_name = f'K{self.sheet}_{const_raw}'
                constants_list.append({'name': go_name, 'value': id_value})

            go_code = template.render(constants=constants_list)
            output_path = paths.SRC_GO_CONSTANTS / f"{self.sheet.lower()}_table_id_constants.go"
            output_path.write_text(go_code, encoding='utf-8')
            logger.info(f"[GO] Generated file: {output_path}")


def process_file(file_path: Path):
    logger.info(f"Start processing file: {file_path}")

    if not file_path.is_file():
        logger.error(f"File not found: {file_path}")
        return

    try:
        converter = ExcelConstantsGenerator(file_path)
        converter.generate_and_save_cpp_constants()
        converter.generate_and_save_go_constants()
        logger.info(f"Finished processing: {file_path}")
    except Exception as e:
        logger.error(f"Error processing file: {file_path}, error: {e}")
    finally:
        if 'converter' in locals():
            converter.close()


def main():
    try:
        paths.SRC_CPP_CONSTANTS.mkdir(parents=True, exist_ok=True)
        paths.SRC_GO_CONSTANTS.mkdir(parents=True, exist_ok=True)

        xlsx_files = utils.get_xlsx_files(paths.DATA_TABLES_DIR)
        if not xlsx_files:
            logger.warning("No Excel files found to process")
            return

        for file in xlsx_files:
            process_file(Path(file))

    except Exception as e:
        logger.error(f"Error during program execution: {e}")
    finally:
        logging.shutdown()


if __name__ == "__main__":
    main()
