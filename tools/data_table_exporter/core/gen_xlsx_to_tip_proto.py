#!/usr/bin/env python
# coding=utf-8

import json
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook
import generate_common  # Assuming generate_common contains the necessary functions
from core import paths
from jinja2 import Environment, FileSystemLoader, select_autoescape

from core.paths import DATA_TABLES_DIR, PROTO_TIP_DIR

# Configure logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')

# === Path setup (using pathlib) ===
excel_file_path = DATA_TABLES_DIR / "tip" / "Tip.xlsx"
json_file_path = paths.GENERATOR_STORAGE_TIP_DIR / "tip_enum_ids.json"

# Ensure output directories exist
PROTO_TIP_DIR.mkdir(parents=True, exist_ok=True)
paths.GENERATOR_STORAGE_TIP_DIR.mkdir(parents=True, exist_ok=True)


def load_existing_ids(json_path: Path):
    """Load existing IDs from a JSON file."""
    if json_path.exists():
        try:
            with json_path.open('r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"Error loading JSON file: {str(e)}")
    return {}


def save_ids_to_json(json_path: Path, data):
    """Save updated IDs to a JSON file."""
    try:
        with json_path.open('w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        logging.info(f"Enum IDs saved to {json_path}")
    except Exception as e:
        logging.error(f"Error saving to JSON file: {str(e)}")


def read_excel_data(file_path: Path, existing_ids):
    """Read data from the first sheet of the provided Excel file path."""
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        num_rows = sheet.max_row

        groups = {}
        current_group = None
        global_row_id = max(
            max((max(data.values()) for data in existing_ids.values()), default=0),
            0
        ) + 1

        for row_idx in range(18, num_rows + 1):
            row_cells = sheet[row_idx]

            if row_cells[0].value and row_cells[0].value.startswith('//'):
                group_name = row_cells[0].value.strip('/').strip()
                current_group = group_name
                groups.setdefault(current_group, {})
            else:
                if current_group:
                    enum_name = row_cells[0].value
                    if enum_name:
                        enum_name = enum_name.strip()
                        if enum_name not in existing_ids.get(current_group, {}):
                            groups[current_group][enum_name] = global_row_id
                            global_row_id += 1
                        else:
                            groups[current_group][enum_name] = existing_ids[current_group][enum_name]

        return groups

    except Exception as e:
        logging.error(f"Error reading Excel file: {str(e)}")
        return {}


def generate_proto_file(group_name, group_data):
    """Generate a Proto file for a given group using Jinja2 template."""
    try:
        template_env = Environment(
            loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'),
            autoescape=select_autoescape(['proto'])
        )
        template = template_env.get_template("tip_enum.proto.j2")

        proto_content = template.render(
            group_name=group_name,
            group_data=group_data
        )

        proto_file_path = PROTO_TIP_DIR / f"{group_name.lower()}_tip.proto"
        proto_file_path.write_text(proto_content, encoding='utf-8')

        logging.info(f"Proto enums file generated: {proto_file_path}")

    except Exception as e:
        logging.error(f"Error generating Proto file for group {group_name}: {str(e)}")


def generate_proto_files(groups):
    """Generate Proto files for all groups using ThreadPoolExecutor."""
    try:
        with ThreadPoolExecutor() as executor:
            futures = [
                executor.submit(generate_proto_file, group_name, group_data)
                for group_name, group_data in groups.items()
            ]

            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    logging.error(f"Task execution failed: {e}")
    except Exception as e:
        logging.error(f"Failed to generate Proto files: {e}")


def main():
    """Main function to read Excel data, generate Proto files, and save JSON."""
    existing_ids = load_existing_ids(json_file_path)
    groups = read_excel_data(excel_file_path, existing_ids)

    if groups:
        generate_proto_files(groups)
        save_ids_to_json(json_file_path, groups)
        logging.info("Proto generation completed.")


if __name__ == "__main__":
    main()
