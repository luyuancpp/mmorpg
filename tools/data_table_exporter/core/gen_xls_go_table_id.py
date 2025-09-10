#!/usr/bin/env python
# coding=utf-8

import openpyxl
import logging
from pathlib import Path

import generate_common  # 你项目已有模块，需包含 BEGIN_ROW_IDX 和 mywrite
from core.constants import DATA_TABLES_DIR  # XLSX_DIR 是 Path 类型

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

GO_DIR = Path("generated/go/table_id/")

def to_pascal_case(s: str) -> str:
    """转为 PascalCase，例如 global_variable -> GlobalVariable"""
    return ''.join(word.capitalize() for word in s.split('_'))

def generate_go_const(sheet):
    """根据 Excel sheet 生成 Go 常量定义"""
    name = sheet.title
    go_struct_name = to_pascal_case(name)

    lines = [f"package table\n", "const ("]
    for row in range(generate_common.BEGIN_ROW_IDX, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1)
        if cell.value is not None:
            const_name = f"{go_struct_name}TableID{int(cell.value)}"
            lines.append(f"    {const_name} = {int(cell.value)}")
    lines.append(")")
    return '\n'.join(lines)

def process_workbook(filepath: Path):
    """读取工作簿并生成 .go 文件"""
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheetname = workbook.sheetnames[0]
        if sheetname not in generate_common.GEN_FILE_LIST:
            return

        sheet = workbook[sheetname]
        go_content = generate_go_const(sheet)
        output_file = GO_DIR / f"{sheetname.lower()}_table_id.go"
        generate_common.mywrite(go_content, str(output_file))
        logger.info(f"Generated Go const file: {output_file}")
    except Exception as e:
        logger.error(f"Failed to process {filepath.name}: {e}")

def main():
    GO_DIR.mkdir(parents=True, exist_ok=True)

    for filepath in DATA_TABLES_DIR.glob("*.xlsx"):
        if filepath.is_file():
            process_workbook(filepath)

if __name__ == "__main__":
    main()
