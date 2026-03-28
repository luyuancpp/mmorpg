#!/usr/bin/env python
# coding=utf-8

import os
import json
import logging
import concurrent.futures
from os import listdir
from os.path import isfile, join
from typing import Any, Union, List, Dict

import openpyxl
import generate_common  # Assumed to contain necessary functions
from core import paths
from paths import PROJECT_GENERATED_JSON_DIR, DATA_TABLES_DIR

# Configure logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def get_column_names(sheet: openpyxl.worksheet.worksheet.Worksheet) -> List[str]:
    """Get column names from first row if fourth row cell is in SERVER_GEN_TYPE"""
    return [
        sheet.cell(row=1, column=col_idx + 1).value
        if sheet.cell(row=4, column=col_idx + 1).value in paths.SERVER_GEN_TYPE
        else ""
        for col_idx in range(sheet.max_column)
    ]


def process_cell_value(cell: openpyxl.cell.cell.Cell, field_type: str) -> Union[float, int, str, Any]:
    """Convert cell value to the specified type"""
    cell_value = cell.value

    if field_type == "string":
        if cell_value in (None, ""):
            return None
        return str(cell_value)

    if field_type in ("float", "double"):
        try:
            return float(cell_value)
        except (ValueError, TypeError):
            return None

    if isinstance(cell_value, float) and cell_value.is_integer():
        return int(cell_value)

    if isinstance(cell_value, int):
        return cell_value

    return cell_value


def handle_map_field_data(cell, row_data: dict, col_name: str, cell_value, map_field_data: dict, column_names: List[str], prev_cell):
    """Handle map-type field data"""
    prev_column_name = column_names[prev_cell.col_idx - 1] if prev_cell else None
    if cell_value in (None, '') and cell.row >= generate_common.BEGIN_ROW_IDX:
        return

    prev_obj_name = generate_common.column_name_to_obj_name(prev_column_name, "_") if prev_column_name else None
    obj_name = generate_common.column_name_to_obj_name(col_name, "_")

    if generate_common.SET_CELL == map_field_data.get(col_name):
        row_data.setdefault(col_name, {})[cell_value] = True
    elif (prev_column_name in map_field_data and
          generate_common.MAP_KEY_CELL == map_field_data.get(prev_column_name) and
          prev_obj_name == obj_name):
        row_data.setdefault(obj_name, {})[prev_cell.value] = cell_value


def handle_array_data(cell, row_data: dict, col_name: str, cell_value):
    """Handle array-type field data"""
    if cell_value in (None, '', 0, -1) and cell.row >= generate_common.BEGIN_ROW_IDX:
        return
    row_data.setdefault(col_name, []).append(cell_value)


def handle_group_data(cell, row_data: dict, col_name: str, cell_value, prev_cell):
    """Handle grouped field data"""
    if cell_value in (None, '', 0, -1) and cell.row >= generate_common.BEGIN_ROW_IDX:
        return

    obj_name = generate_common.column_name_to_obj_name(col_name, "_")
    member_dict = {col_name: cell_value}

    if obj_name in row_data:
        last_element = row_data[obj_name][-1]
        if col_name in last_element:
            row_data[obj_name].append(member_dict)
        else:
            last_element[col_name] = cell_value
    else:
        row_data[obj_name] = [member_dict]


def process_row(sheet, row, column_names: List[str]) -> dict:
    """Process a single row of data"""
    sheet_data = generate_common.get_sheet_data(sheet, column_names)
    array_data = sheet_data[generate_common.SHEET_ARRAY_DATA_INDEX]
    field_type_data = sheet_data[generate_common.FILE_TYPE_INDEX]
    group_data = sheet_data[generate_common.SHEET_GROUP_ARRAY_DATA_INDEX]
    map_field_data = sheet_data[generate_common.MAP_TYPE_INDEX]

    row_data = {}
    prev_cell = None

    for idx, cell in enumerate(row):
        if idx >= len(column_names):
            logger.warning(f"Row {cell.row} in sheet '{sheet.title}' has more cells than column names.")
            break

        col_name = column_names[idx]
        if not col_name.strip():
            continue

        cell_value = process_cell_value(cell, field_type_data.get(col_name, ""))
        if cell_value is None:
            continue

        if col_name in map_field_data or generate_common.is_key_in_map(group_data, col_name, map_field_data, column_names):
            handle_map_field_data(cell, row_data, col_name, cell_value, map_field_data, column_names, prev_cell)
        elif col_name in array_data:
            handle_array_data(cell, row_data, col_name, cell_value)
        elif generate_common.is_key_in_group_array(group_data, col_name, column_names):
            handle_group_data(cell, row_data, col_name, cell_value, prev_cell)
        else:
            if cell_value in (None, '') and cell.row >= generate_common.BEGIN_ROW_IDX:
                logger.error(f"Sheet '{sheet.title}', Cell {cell.coordinate} is empty or invalid.")
            row_data[col_name] = cell_value

        prev_cell = cell

    return row_data


def extract_sheet_data(sheet, column_names: List[str]) -> List[dict]:
    """Extract data from entire sheet"""
    return [
        process_row(sheet, row, column_names)
        for row in sheet.iter_rows(min_row=generate_common.BEGIN_ROW_IDX + 1, values_only=False)
    ]


def extract_workbook_data(workbook: openpyxl.Workbook) -> Dict[str, List[dict]]:
    """Extract data from the workbook; only the first sheet is used"""
    data = {}
    if workbook.sheetnames:
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        if sheet.cell(row=1, column=1).value != "id":
            logger.error(f"The first column of sheet '{sheet_name}' must be 'id'.")
        else:
            column_names = get_column_names(sheet)
            data[sheet_name] = extract_sheet_data(sheet, column_names)
    else:
        logger.error("No sheets found in the workbook.")
    return data


def save_json(data: dict, file_path: str) -> None:
    """Save data to a JSON file"""
    json_data = json.dumps({"data": data}, sort_keys=True, indent=1, separators=(',', ': '))
    json_data = json_data.replace('"[', '[').replace(']"', ']')
    json_data = json_data.replace('\r\n', '\n').replace('\r', '\n')

    try:
        with open(file_path, 'w', encoding='utf-8', newline='') as f:
            f.write(json_data)
            logger.info(f"Generated JSON file: {file_path}")
    except IOError as e:
        logger.error(f"Failed to save JSON file {file_path}: {e}")


def process_excel_file(file_path: str) -> None:
    """Process a single Excel file and generate JSON"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        workbook_data = extract_workbook_data(workbook)

        for sheet_name, data in workbook_data.items():
            json_file_path = os.path.join(PROJECT_GENERATED_JSON_DIR, f"{sheet_name}.json")
            save_json(data, json_file_path)

    except Exception as e:
        logger.error(f"Failed to process file {file_path}: {e}")


def main() -> None:
    """Main function to process all Excel files in the specified directory"""
    os.makedirs(PROJECT_GENERATED_JSON_DIR, exist_ok=True)

    files = [
        join(DATA_TABLES_DIR, filename)
        for filename in listdir(DATA_TABLES_DIR)
        if isfile(join(DATA_TABLES_DIR, filename)) and filename.endswith('.xlsx')
    ]

    max_workers = os.cpu_count() or 1
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        executor.map(process_excel_file, files)


if __name__ == "__main__":
    main()
