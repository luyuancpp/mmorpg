#!/usr/bin/env python
# coding=utf-8

import json
import logging
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook  # Use openpyxl for better support of .xlsx files

from core.paths import (
    PROJECT_GENERATED_CODE_PROTO_OPERATOR_DIR,
    GENERATOR_STORAGE_OPERATOR_DIR,
    PROJECT_OPERATOR_XLSX,
    GENERATOR_STORAGE_OPERATOR_FILE_DIR
)

# Configure logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')

# Ensure output directories exist
GENERATOR_STORAGE_OPERATOR_DIR.mkdir(parents=True, exist_ok=True)
PROJECT_GENERATED_CODE_PROTO_OPERATOR_DIR.mkdir(parents=True, exist_ok=True)

# Use a lock for thread-safe file access
file_lock = threading.Lock()

def read_temp_id_mapping():
    """Reads existing ID mappings from temp file."""
    try:
        if GENERATOR_STORAGE_OPERATOR_FILE_DIR.exists():
            with GENERATOR_STORAGE_OPERATOR_FILE_DIR.open('r', encoding='utf-8') as f:
                return json.load(f)
        else:
            return {}
    except Exception as e:
        logging.error(f"Error reading ID mapping file: {str(e)}")
        return {}


def write_temp_id_mapping(mapping):
    """Writes ID mappings to temp file."""
    try:
        with file_lock:
            with GENERATOR_STORAGE_OPERATOR_FILE_DIR.open('w', encoding='utf-8') as f:
                json.dump(mapping, f, indent=2)
    except Exception as e:
        logging.error(f"Error writing ID mapping file: {str(e)}")


def read_excel_data(file_path: Path):
    """Reads data from the first sheet of the provided Excel file."""
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active  # Only read the first sheet
        num_rows = sheet.max_row

        groups = {}
        current_group = None
        current_group_data = []
        global_row_id = 1

        for row_idx in range(18, num_rows + 1):  # Starting from row 18
            row_cells = sheet[row_idx]

            if row_cells[0].value and row_cells[0].value.startswith('//'):
                group_name = row_cells[0].value.strip('/').strip()

                if current_group:
                    groups[current_group] = current_group_data

                current_group = group_name
                current_group_data = []
            else:
                if current_group:
                    enum_name = row_cells[0].value
                    if enum_name:
                        current_group_data.append((enum_name.strip(), global_row_id))
                        global_row_id += 1

        if current_group:
            groups[current_group] = current_group_data

        return groups

    except Exception as e:
        logging.error(f"Error reading Excel file: {str(e)}")
        return {}


def generate_proto_file(group_name, group_data, existing_id_mapping):
    """Generates Proto file for a given group."""
    try:
        proto_content = 'syntax = "proto3";\n\n'
        proto_content += 'option go_package = "generated/pb/table";\n\n'
        proto_content += f"enum {group_name} {{\n"
        proto_content += f"  k{group_name.capitalize()}OK = 0;\n"

        for enum_name, _ in group_data:
            if enum_name in existing_id_mapping:
                enum_id = existing_id_mapping[enum_name]
            else:
                enum_id = max(existing_id_mapping.values(), default=0) + 1
                existing_id_mapping[enum_name] = enum_id

            enum_name_with_k = f"k{enum_name.strip()}"
            proto_content += f"  {enum_name_with_k} = {enum_id};\n"

        proto_content += '};\n'

        proto_file_path = PROJECT_GENERATED_CODE_PROTO_OPERATOR_DIR / f"{group_name.lower()}_operator.proto"

        with file_lock:
            proto_file_path.write_text(proto_content, encoding='utf-8')

        logging.info(f"Proto enums file generated: {proto_file_path}")

        # Write updated ID mapping
        write_temp_id_mapping(existing_id_mapping)

    except Exception as e:
        logging.error(f"Error generating Proto file for group {group_name}: {str(e)}")


def generate_proto_files(groups):
    """Generates Proto files for all groups using ThreadPoolExecutor."""
    existing_id_mapping = read_temp_id_mapping()

    with ThreadPoolExecutor() as executor:
        futures = [
            executor.submit(generate_proto_file, group_name, group_data, existing_id_mapping)
            for group_name, group_data in groups.items()
        ]

        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logging.error(f"Error occurred: {str(e)}")


def main():
    """Main function to read Excel data and generate Proto files."""
    groups = read_excel_data(PROJECT_OPERATOR_XLSX)
    if groups:
        generate_proto_files(groups)
        logging.info("Proto generation completed.")


if __name__ == "__main__":
    main()
