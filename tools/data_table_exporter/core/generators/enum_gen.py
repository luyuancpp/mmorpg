"""Enum generators for Operator and Tip Excel files.

* **Operator.xlsx** → one ``.proto`` per group with an ``enum`` definition.
* **Tip.xlsx**      → one ``.proto`` per error group with an ``enum`` definition.

Both use persistent ID mappings in ``state/`` so that enum values are stable
across regeneration.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

from core.config_loader import ExporterConfig
from core.file_utils import ensure_dirs, write_file

logger = logging.getLogger(__name__)


# ===================================================================
# Operator enum generation
# ===================================================================

def generate_operator_enums(cfg: ExporterConfig) -> None:
    """Read ``Operator.xlsx`` and produce one proto file per group."""
    if not cfg.operator_file.exists():
        logger.warning("Operator file not found: %s", cfg.operator_file)
        return

    state_dir = cfg.state_dir / "operator"
    ensure_dirs(state_dir)
    id_file = state_dir / "id_pool.json"
    id_map = _load_json(id_file)

    groups = _read_operator_groups(cfg.operator_file)
    out_dir = cfg.proto_dir / "operator"
    ensure_dirs(out_dir)

    env = Environment(loader=FileSystemLoader(str(cfg.template_dir), encoding="utf-8"))
    tpl = env.get_template("operator_enum.proto.j2")

    for group_name, entries in groups.items():
        for enum_name, _ in entries:
            if enum_name not in id_map:
                id_map[enum_name] = max(id_map.values(), default=0) + 1

        content = tpl.render(group_name=group_name, entries=entries, id_map=id_map)
        write_file(out_dir / f"{group_name.lower()}_operator.proto", content)
        logger.info("Generated operator proto: %s", group_name)

    _save_json(id_file, id_map)


def _read_operator_groups(file_path: Path) -> dict[str, list[tuple[str, int]]]:
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    groups: dict[str, list[tuple[str, int]]] = {}
    current_group = None
    row_id = 1

    for row_idx in range(18, (ws.max_row or 18) + 1):
        cells = ws[row_idx]
        val = cells[0].value
        if val and str(val).startswith("//"):
            current_group = str(val).strip("/").strip()
            groups.setdefault(current_group, [])
        elif current_group and val:
            groups[current_group].append((str(val).strip(), row_id))
            row_id += 1

    return groups


# ===================================================================
# Tip enum generation
# ===================================================================

def generate_tip_enums(cfg: ExporterConfig) -> None:
    """Read ``Tip.xlsx`` and produce one proto file per error group."""
    tip_file = cfg.tip_file
    if not tip_file.exists():
        logger.warning("Tip file not found: %s", tip_file)
        return

    state_dir = cfg.state_dir / "mapping" / "tip_enum_ids"
    ensure_dirs(state_dir)
    id_file = state_dir / "tip_enum_ids.json"
    existing_ids = _load_json(id_file)

    groups = _read_tip_groups(tip_file, existing_ids)
    out_dir = cfg.proto_dir / "tip"
    ensure_dirs(out_dir)

    env = Environment(loader=FileSystemLoader(str(cfg.template_dir), encoding="utf-8"))
    tpl = env.get_template("tip_enum.proto.j2")

    for group_name, group_data in groups.items():
        content = tpl.render(group_name=group_name, group_data=group_data)
        write_file(out_dir / f"{group_name.lower()}_tip.proto", content)
        logger.info("Generated tip proto: %s", group_name)

    _save_json(id_file, groups)


def _read_tip_groups(file_path: Path, existing_ids: dict) -> dict[str, dict[str, int]]:
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    groups: dict[str, dict[str, int]] = {}
    current_group = None

    global_id = 1
    for v in existing_ids.values():
        if isinstance(v, dict):
            for num in v.values():
                global_id = max(global_id, int(num) + 1)

    for row_idx in range(18, (ws.max_row or 18) + 1):
        cells = ws[row_idx]
        val = cells[0].value
        if val and str(val).startswith("//"):
            current_group = str(val).strip("/").strip()
            groups.setdefault(current_group, {})
        elif current_group and val:
            enum_name = str(val).strip()
            if enum_name in existing_ids.get(current_group, {}):
                groups[current_group][enum_name] = existing_ids[current_group][enum_name]
            else:
                groups[current_group][enum_name] = global_id
                global_id += 1

    return groups


# ===================================================================
# Shared helpers
# ===================================================================

def _load_json(path: Path) -> dict:
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as exc:
            logger.error("Failed to load %s: %s", path, exc)
    return {}


def _save_json(path: Path, data: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
