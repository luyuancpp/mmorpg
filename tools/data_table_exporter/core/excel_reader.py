"""Excel reader — parse ``.xlsx`` metadata into :class:`TableSchema` objects.

Also provides data-access helpers so that generators never need to
re-open or re-parse workbooks themselves.

Excel header format (5 rows):
    Row 1: field name — one name per logical field span (first col of span)
    Row 2: type declaration — ``uint32``, ``map<K,V>``, ``set<T>``,
           ``repeated uint32``, ``repeated { uint32 f1; uint32 f2 }``
    Row 3: owner — server | client | common | design (first col of span)
    Row 4: options — space-separated tokens (first col of span)
    Row 5: comment (first col of span)
    Row 6+: data
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from core.config_loader import ExporterConfig
from core.file_utils import list_xlsx
from core.schema import (
    ArrayField,
    ColumnDef,
    GroupField,
    MapField,
    TableSchema,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Public API — schema reading
# ---------------------------------------------------------------------------

def read_table(file_path: Path, cfg: ExporterConfig) -> Optional[TableSchema]:
    """Read a single ``.xlsx`` and return its :class:`TableSchema`."""
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as exc:
        logger.error("Cannot open %s: %s", file_path, exc)
        return None

    if not wb.sheetnames:
        logger.error("No sheets in %s", file_path)
        return None

    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    first_cell = str(ws.cell(row=1, column=1).value or "").strip()
    if first_cell != "id":
        logger.error("First column name must be 'id' (%s)", file_path)
        return None

    columns = _parse_columns(ws, cfg)
    arrays, groups, maps = _detect_layout(columns)

    first_col = columns[0] if columns else None
    use_flat_multimap = first_col is not None and first_col.is_multi_key

    constants_idx = next(
        (c.excel_index for c in columns if c.name == "constants_name"), None
    )

    return TableSchema(
        name=sheet_name,
        source_path=file_path,
        columns=columns,
        arrays=arrays,
        groups=groups,
        maps=maps,
        use_flat_multimap=use_flat_multimap,
        has_constants_name=constants_idx is not None,
        constants_name_index=constants_idx,
    )


def read_all_tables(cfg: ExporterConfig) -> list[TableSchema]:
    """Read every ``.xlsx`` in *cfg.data_dir* and return a list of schemas."""
    tables: list[TableSchema] = []
    for xlsx in list_xlsx(cfg.data_dir):
        schema = read_table(xlsx, cfg)
        if schema is not None:
            tables.append(schema)
    return tables


# ---------------------------------------------------------------------------
# Public API — data access helpers (for generators)
# ---------------------------------------------------------------------------

def open_worksheet(schema: TableSchema) -> Worksheet:
    """Open the source workbook and return the first worksheet."""
    wb = openpyxl.load_workbook(schema.source_path)
    return wb[schema.name]


def read_id_column(schema: TableSchema, cfg: ExporterConfig) -> list[int]:
    """Read all ``id`` values (first column) from data rows."""
    ws = open_worksheet(schema)
    return [
        int(row[0])
        for row in ws.iter_rows(min_row=cfg.data_begin_row, values_only=True)
        if row[0] is not None
    ]


def read_data_rows(schema: TableSchema, cfg: ExporterConfig) -> list[dict]:
    """Read data rows and return them as dicts with structural awareness.

    Arrays become lists, groups become nested dicts, maps become dicts,
    sets become ``{value: True}`` dicts.
    """
    ws = open_worksheet(schema)
    server_names = {c.name for c in schema.columns if c.is_server}
    col_names = [c.name for c in schema.columns]
    col_types = {c.name: c.data_type for c in schema.columns}
    col_to_group = schema.col_to_group

    rows: list[dict] = []
    for excel_row in ws.iter_rows(min_row=cfg.data_begin_row, values_only=False):
        row_data = _build_row(excel_row, schema, col_names, col_types,
                              server_names, col_to_group)
        if row_data:
            rows.append(row_data)
    return rows


# ---------------------------------------------------------------------------
# Internal — column parsing
# ---------------------------------------------------------------------------

def _cell_str(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


# ---------------------------------------------------------------------------
# Internal — declaration parsing (proto-style Row 1)
# ---------------------------------------------------------------------------

# Regex patterns for type declarations (name-free)
_RE_MAP = re.compile(r'^map\s*<\s*(\w+)\s*,\s*(\w+)\s*>$')
_RE_SET = re.compile(r'^set\s*<\s*(\w+)\s*>$')
_RE_REPEATED_MSG = re.compile(r'^repeated\s*\{(.+)\}$')
_RE_REPEATED = re.compile(r'^repeated\s+(\w+)$')
_RE_SCALAR = re.compile(r'^(\w+)$')


def _parse_declaration(type_text: str, name: str):
    """Parse a type declaration string together with a field name.

    Returns a tuple whose first element is the kind:
        ('scalar', name, data_type)
        ('map', name, key_type, value_type)
        ('set', name, value_type)
        ('repeated', name, data_type)
        ('message', name, [(data_type, field_name), ...])
    """
    text = type_text.strip()

    m = _RE_MAP.match(text)
    if m:
        return ('map', name, m.group(1), m.group(2))

    m = _RE_SET.match(text)
    if m:
        return ('set', name, m.group(1))

    m = _RE_REPEATED_MSG.match(text)
    if m:
        body = m.group(1)
        sub_fields = []
        for part in body.split(';'):
            part = part.strip()
            if not part:
                continue
            tokens = part.split()
            if len(tokens) == 2:
                sub_fields.append((tokens[0], tokens[1]))
            else:
                raise ValueError(f"Cannot parse message sub-field: '{part}'")
        return ('message', name, sub_fields)

    m = _RE_REPEATED.match(text)
    if m:
        return ('repeated', name, m.group(1))

    m = _RE_SCALAR.match(text)
    if m:
        return ('scalar', name, m.group(1))

    raise ValueError(f"Cannot parse type declaration: '{text}'")


# ---------------------------------------------------------------------------
# Internal — column parsing (proto-style header)
# ---------------------------------------------------------------------------

def _parse_columns(ws, cfg: ExporterConfig) -> list[ColumnDef]:
    """Parse Row 1 (name) + Row 2 (type) into a flat list of ColumnDefs."""
    meta = cfg.metadata_rows
    max_col = ws.max_column or 0

    # Row numbers
    name_row = meta.get("field_name", 1)
    type_row = meta.get("field_type", 2)
    owner_row = meta.get("owner", 3)
    options_row = meta.get("options", 4)
    comment_row = meta.get("comment", 5)

    # Step 1: find non-empty cells in the name row (field boundaries)
    name_starts: list[tuple[int, str]] = []      # (0-based col index, name)
    for idx in range(max_col):
        text = _cell_str(ws, name_row, idx + 1)
        if text:
            name_starts.append((idx, text))

    if not name_starts:
        return []

    # Step 2: compute spans — each field runs until the next name
    spans: list[tuple[int, int, str]] = []        # (start, end_exclusive, name)
    for i, (start, name) in enumerate(name_starts):
        end = name_starts[i + 1][0] if i + 1 < len(name_starts) else max_col
        spans.append((start, end, name))

    # Step 3: parse each type declaration and expand into ColumnDefs
    columns: list[ColumnDef] = []
    for start, end, field_name in spans:
        span_size = end - start

        type_text = _cell_str(ws, type_row, start + 1)
        owner = _cell_str(ws, owner_row, start + 1).lower()
        options_raw = _cell_str(ws, options_row, start + 1)
        comment = _cell_str(ws, comment_row, start + 1)
        options = options_raw.split() if options_raw else []

        parsed = _parse_declaration(type_text, field_name)
        kind = parsed[0]

        if kind == 'scalar':
            _, name, data_type = parsed
            if span_size > 1:
                # Auto-detect repeated: scalar spanning multiple columns
                for i in range(span_size):
                    columns.append(ColumnDef(
                        name=name, data_type=data_type, owner=owner,
                        struct='repeated',
                        options=options if i == 0 else [],
                        comment=comment if i == 0 else '',
                        excel_index=start + i,
                    ))
            else:
                columns.append(ColumnDef(
                    name=name, data_type=data_type, owner=owner,
                    struct='', options=options, comment=comment,
                    excel_index=start,
                ))

        elif kind == 'map':
            _, name, key_type, value_type = parsed
            for i in range(0, span_size, 2):
                columns.append(ColumnDef(
                    name=f'{name}_key', data_type=key_type, owner=owner,
                    struct='map_key', options=options if i == 0 else [],
                    comment=comment if i == 0 else '',
                    excel_index=start + i,
                ))
                if start + i + 1 < end:
                    columns.append(ColumnDef(
                        name=f'{name}_value', data_type=value_type, owner=owner,
                        struct='map_value', options=[],
                        comment='', excel_index=start + i + 1,
                    ))

        elif kind == 'set':
            _, name, value_type = parsed
            for i in range(span_size):
                columns.append(ColumnDef(
                    name=name, data_type=value_type, owner=owner,
                    struct='set', options=options if i == 0 else [],
                    comment=comment if i == 0 else '',
                    excel_index=start + i,
                ))

        elif kind == 'repeated':
            _, name, data_type = parsed
            for i in range(span_size):
                columns.append(ColumnDef(
                    name=name, data_type=data_type, owner=owner,
                    struct='repeated', options=options if i == 0 else [],
                    comment=comment if i == 0 else '',
                    excel_index=start + i,
                ))

        elif kind == 'message':
            _, msg_name, sub_fields = parsed
            n_sub = len(sub_fields)
            for i in range(span_size):
                sf_type, sf_name = sub_fields[i % n_sub]
                columns.append(ColumnDef(
                    name=f'{msg_name}_{sf_name}', data_type=sf_type, owner=owner,
                    struct=f'message:{msg_name}',
                    options=options if i == 0 else [],
                    comment=comment if i == 0 else '',
                    excel_index=start + i,
                ))

    return columns


# ---------------------------------------------------------------------------
# Internal — layout detection (driven by struct row)
# ---------------------------------------------------------------------------

def _detect_layout(
    columns: list[ColumnDef],
) -> tuple[dict[str, ArrayField], dict[str, GroupField], dict[str, MapField]]:
    """Detect arrays, message groups, and maps from the struct row."""

    # --- repeated fields ---
    arrays: dict[str, ArrayField] = {}
    repeated_cols: dict[str, list[ColumnDef]] = {}
    for col in columns:
        if col.struct == "repeated":
            repeated_cols.setdefault(col.name, []).append(col)
    for name, cols in repeated_cols.items():
        arrays[name] = ArrayField(
            name=name,
            data_type=cols[0].data_type,
            indices=[c.excel_index for c in cols],
        )

    # --- message groups ---
    groups: dict[str, GroupField] = {}
    for col in columns:
        if col.struct.startswith("message:"):
            msg_name = col.struct.split(":", 1)[1]
            if msg_name not in groups:
                groups[msg_name] = GroupField(name=msg_name, columns=[], indices=[])
            # Only add unique column names to columns (for sub-message definition)
            existing = {c.name for c in groups[msg_name].columns}
            if col.name not in existing:
                groups[msg_name].columns.append(col)
            groups[msg_name].indices.append(col.excel_index)

    # --- maps (consecutive map_key + map_value pairs) ---
    maps: dict[str, MapField] = {}
    i = 0
    while i < len(columns):
        if columns[i].struct == "map_key" and i + 1 < len(columns) and columns[i + 1].struct == "map_value":
            prefix = _common_prefix(columns[i].name, columns[i + 1].name)
            if prefix and prefix not in maps:
                maps[prefix] = MapField(
                    name=prefix,
                    key_type=columns[i].data_type,
                    value_type=columns[i + 1].data_type,
                )
            i += 2
        else:
            i += 1

    return arrays, groups, maps


def _common_prefix(a: str, b: str) -> str:
    """Find common underscore-delimited prefix."""
    parts_a, parts_b = a.split("_"), b.split("_")
    common = []
    for pa, pb in zip(parts_a, parts_b):
        if pa == pb:
            common.append(pa)
        else:
            break
    return "_".join(common)


# ---------------------------------------------------------------------------
# Internal — data row building
# ---------------------------------------------------------------------------

def _build_row(
    excel_row,
    schema: TableSchema,
    col_names: list[str],
    col_types: dict[str, str],
    server_names: set[str],
    col_to_group: dict[int, str],
) -> dict:
    """Convert one Excel data row into a structured dict."""
    row_data: dict = {}
    pending_map_key = None

    for idx, cell in enumerate(excel_row):
        if idx >= len(col_names):
            break
        name = col_names[idx]
        if not name or name not in server_names:
            continue

        value = _convert_cell(cell.value, col_types.get(name, ""))
        if value is None:
            pending_map_key = None
            continue

        col = schema.columns[idx]

        if col.map_role == "set":
            row_data.setdefault(name, {})[value] = True
        elif col.map_role == "map_key":
            pending_map_key = value
        elif col.map_role == "map_value" and pending_map_key is not None:
            prefix = _common_prefix(col_names[idx - 1], name) if idx > 0 else name
            row_data.setdefault(prefix, {})[pending_map_key] = value
            pending_map_key = None
        elif name in schema.arrays:
            if value not in (0, -1, ""):
                row_data.setdefault(name, []).append(value)
        elif idx in col_to_group:
            group_name = col_to_group[idx]
            _append_to_group(row_data, group_name, name, value)
        else:
            row_data[name] = value

    return row_data


def _append_to_group(row_data: dict, group_name: str, col_name: str, value) -> None:
    """Append a value into the correct group sub-list."""
    member = {col_name: value}
    if group_name in row_data:
        last = row_data[group_name][-1]
        if col_name in last:
            row_data[group_name].append(member)
        else:
            last[col_name] = value
    else:
        row_data[group_name] = [member]


def _convert_cell(value, data_type: str):
    """Convert an Excel cell value to the appropriate Python type."""
    if value is None or value == "":
        return None
    if data_type == "string":
        return str(value)
    if data_type in ("float", "double"):
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value
