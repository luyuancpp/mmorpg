#!/usr/bin/env python
# coding=utf-8

import concurrent.futures
import openpyxl
import logging
from pathlib import Path
from multiprocessing import cpu_count
from jinja2 import Environment, FileSystemLoader

import generate_common  # 你项目中的工具模块
from core.paths import SRC_CPP, PROJECT_GENERATED_CODE_TABLE_GO_DIR, DATA_TABLES_DIR

# 日志配置
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


# 类型转换函数
def get_cpp_type_name(type_name):
    if type_name == 'string':
        return 'std::string'
    elif 'int' in type_name:
        return f'{type_name}_t'
    return type_name


def get_cpp_type_param_name_with_ref(type_name):
    if type_name == 'string':
        return 'const std::string&'
    elif 'int' in type_name:
        return f'{type_name}_t'
    return type_name


def convert_to_go_type(col_type):
    if "int32" in col_type:
        return "int32"
    elif "int64" in col_type:
        return "int64"
    elif "float" in col_type:
        return "float32"
    elif "double" in col_type:
        return "float64"
    elif "bool" in col_type:
        return "bool"
    elif "string" in col_type:
        return "string"
    return "interface{}"


def get_workbook_data(workbook):
    """提取工作簿中第一张 sheet 的必要元数据"""
    workbook_data = {}
    sheet_names = workbook.sheetnames
    if sheet_names:
        sheet = workbook[sheet_names[0]]
        cell_value = sheet['A5'].value
        use_flat_multimap = cell_value is not None and cell_value.lower() == 'multi'
        first_19_rows_per_column = generate_common.get_first_19_rows_per_column(sheet)
        workbook_data[sheet_names[0]] = {
            'multi': use_flat_multimap,
            'get_first_19_rows_per_column': first_19_rows_per_column
        }
    return workbook_data


def process_workbook(filepath: Path):
    """处理单个 xlsx 文件生成代码"""
    try:
        workbook = openpyxl.load_workbook(filepath)
        workbook_data = get_workbook_data(workbook)
        for sheetname, data in workbook_data.items():
            sheetname_lower = sheetname.lower()
            env = Environment(loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'), auto_reload=True)

            # === C++ Header ===
            header_template = env.get_template('config_template.h.jinja')
            header_content = header_template.render(
                datastring=data['get_first_19_rows_per_column'],
                sheetname=sheetname,
                use_flat_multimap=data['multi'],
                generate_common=generate_common,
                get_cpp_type_param_name_with_ref=get_cpp_type_param_name_with_ref,
                get_cpp_type_name=get_cpp_type_name
            )
            generate_common.mywrite(header_content, SRC_CPP / f"{sheetname_lower}_table.h")

            # === C++ Implementation ===
            cpp_template = env.get_template('config_template.cpp.jinja')
            cpp_content = cpp_template.render(
                datastring=data['get_first_19_rows_per_column'],
                sheetname=sheetname,
                generate_common=generate_common,
                get_cpp_type_param_name_with_ref=get_cpp_type_param_name_with_ref,
                get_cpp_type_name=get_cpp_type_name
            )
            generate_common.mywrite(cpp_content, SRC_CPP / f"{sheetname_lower}_table.cpp")

            # === Go Code ===
            go_template = env.get_template("config_template.go.jinja")
            go_content = go_template.render(
                datastring=data['get_first_19_rows_per_column'],
                sheetname=sheetname,
                generate_common=generate_common,
                convert_to_go_type=convert_to_go_type,
                proto_import_path="your/proto/package/path"  # ⚠️ 修改为你的真实路径
            )
            PROJECT_GENERATED_CODE_TABLE_GO_DIR.mkdir(parents=True, exist_ok=True)
            generate_common.mywrite(go_content, PROJECT_GENERATED_CODE_TABLE_GO_DIR / f"{sheetname_lower}_table.go")

    except Exception as e:
        logging.error(f"Failed to load or process workbook {filepath.name}: {e}")


def generate_all_config():
    """生成 all_table.h / all_table.cpp / all_table.go 等聚合文件"""
    sheetnames = set()
    xlsx_files = sorted(DATA_TABLES_DIR.glob("*.xlsx"), key=lambda f: f.stat().st_size, reverse=True)
    for filepath in xlsx_files:
        try:
            workbook = openpyxl.load_workbook(filepath)
            workbook_data = get_workbook_data(workbook)
            sheetnames.update(workbook_data.keys())
        except Exception as e:
            logging.error(f"Failed to process file {filepath}: {e}")

    sheetnames = sorted(sheetnames)
    cpucount = cpu_count()

    env = Environment(loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'))
    header_template = env.get_template("all_table.h.jinja")
    cpp_template = env.get_template("all_table.cpp.jinja")
    go_template = env.get_template("all_table.go.jinja")

    header_content = header_template.render()
    cpp_content = cpp_template.render(sheetnames=sheetnames, cpucount=cpucount)
    go_content = go_template.render(sheetnames=sheetnames)

    PROJECT_GENERATED_CODE_TABLE_GO_DIR.mkdir(parents=True, exist_ok=True)
    generate_common.mywrite(go_content, PROJECT_GENERATED_CODE_TABLE_GO_DIR / "all_table.go")

    return header_content, cpp_content


def main(XLS_DIR: Path = None):
    """主函数，支持传入自定义 Excel 目录"""
    xls_dir = XLS_DIR or DATA_TABLES_DIR
    SRC_CPP.mkdir(parents=True, exist_ok=True)
    PROJECT_GENERATED_CODE_TABLE_GO_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_files = list(xls_dir.glob("*.xlsx"))
    with concurrent.futures.ProcessPoolExecutor() as executor:
        executor.map(process_workbook, xlsx_files)

    header_content, cpp_content = generate_all_config()
    generate_common.mywrite(header_content, SRC_CPP / "all_table.h")
    generate_common.mywrite(cpp_content, SRC_CPP / "all_table.cpp")


if __name__ == "__main__":
    main()
