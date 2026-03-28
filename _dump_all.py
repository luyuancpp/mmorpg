import openpyxl
import os

data_dir = r"data"
for fname in sorted(os.listdir(data_dir)):
    if not fname.endswith(".xlsx"):
        continue
    fpath = os.path.join(data_dir, fname)
    print()
    print("=" * 80)
    print("FILE:", fname)
    print("=" * 80)
    wb = openpyxl.load_workbook(fpath)
    ws = wb[wb.sheetnames[0]]
    print("Sheet:", wb.sheetnames[0], " Rows:", ws.max_row, " Cols:", ws.max_column)
    for r in range(1, min(22, (ws.max_row or 1) + 1)):
        cells = []
        for c in range(1, min((ws.max_column or 0) + 1, 30)):
            v = ws.cell(row=r, column=c).value
            s = str(v)[:25] if v is not None else ""
            cells.append(s)
        # Mark which rows have non-empty cells (beyond first col)
        nonempty = sum(1 for c in cells[1:] if c)
        mark = "*" if nonempty > 0 else " "
        print("  R%02d %s: %s" % (r, mark, " | ".join(cells)))
