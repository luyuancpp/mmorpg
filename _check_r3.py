import openpyxl
wb = openpyxl.load_workbook('data/Test.xlsx')
ws = wb[wb.sheetnames[0]]
for r in range(1, 4):
    vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        vals.append(repr(v))
    print(f"R{r:02d}: {vals}")
