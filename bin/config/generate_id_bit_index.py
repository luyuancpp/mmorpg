#!/usr/bin/env python
# coding=utf-8

import os
import json
import logging
import concurrent.futures
from os import listdir
from os.path import isfile, join
import openpyxl
import multiprocessing
from typing import List, Optional, Dict

import gen_common  # Assuming gen_common contains the necessary functions
from common import constants

# Setup Logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelToCppConverter:
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.workbook = openpyxl.load_workbook(excel_file)
        self.sheet = self.workbook.sheetnames[0]
        self.worksheet = self.workbook[self.sheet]
        self.bit_index_col = self._find_bit_index_column()
        self.mapping_file = constants.GENERATOR_CONSTANTS_NAME_DIR + self.sheet.lower() + '_mapping.json'

    def _find_bit_index_column(self) -> Optional[int]:
        """Find the index of the 'bit_index' column."""
        headers = [cell.value for cell in self.worksheet[1]]
        return headers.index('bit_index') if 'bit_index' in headers else None

    def _load_existing_mapping(self) -> Dict[int, int]:
        """Load existing ID to index mapping from a JSON file."""
        if os.path.exists(self.mapping_file):
            with open(self.mapping_file, 'r') as file:
                return json.load(file)
        return {}

    def _save_mapping(self, mapping: Dict[int, int]) -> None:
        """Save the ID to index mapping to a JSON file."""
        with open(self.mapping_file, 'w') as file:
            json.dump(mapping, file, indent=4)

    def should_process(self) -> bool:
        """Check if the worksheet contains the 'bit_index' column."""
        return self.bit_index_col is not None

    def generate_cpp_constants(self) -> str:
        """Generate C++ constants from the Excel data."""
        cpp_constants = "#pragma once\n\n"
        id_to_index = self._load_existing_mapping()
        current_index = max(id_to_index.values(), default=-1) + 1

        for row in self.worksheet.iter_rows(min_row=20, values_only=True):
            id_value = row[0]
            if id_value is None:
                continue  # Skip rows with no ID

            if id_value not in id_to_index:
                id_to_index[id_value] = current_index
                current_index += 1

        for row in self.worksheet.iter_rows(min_row=20, values_only=True):
            id_value = row[0]
            if id_value is None:
                continue  # Skip rows with no ID

            constant_name = self._generate_constant_name(row, id_to_index[id_value])
            cpp_constant = f"constexpr uint32_t {constant_name} = {id_to_index[id_value]};\n"
            cpp_constants += cpp_constant

        self._save_mapping(id_to_index)
        return cpp_constants

    def _generate_constant_name(self, row: tuple, index: int) -> str:
        """Generate the constant name based on the 'bit_index' or ID value."""
        if self.bit_index_col is not None and row[self.bit_index_col]:
            return f'k{self.sheet}_{row[self.bit_index_col]}'
        return f"k{self.sheet}_{index}"

    def save_cpp_constants_to_file(self, cpp_constants: str) -> None:
        """Save the generated C++ constants to a file."""
        output_file = constants.GENERATOR_CONSTANTS_NAME_DIR + self.sheet.lower() + '_table_id_bit_index.h'
        with open(output_file, 'w') as file:
            file.write(cpp_constants)

def get_xlsx_files(directory: str) -> List[str]:
    """List all .xlsx files in the specified directory."""
    return [join(directory, filename) for filename in listdir(directory)
            if isfile(join(directory, filename)) and filename.endswith('.xlsx')]

def process_file(excel_file: str) -> None:
    """Process each Excel file and generate the corresponding C++ constants file."""
    converter = ExcelToCppConverter(excel_file)
    if converter.should_process():
        cpp_constants = converter.generate_cpp_constants()
        converter.save_cpp_constants_to_file(cpp_constants)
    else:
        logger.info(f"Skipping file {excel_file} as it does not contain 'bit_index' column.")

def main() -> None:
    """Main function to process all Excel files."""
    os.makedirs(constants.GENERATOR_CONSTANTS_NAME_DIR, exist_ok=True)

    try:
        xlsx_files = get_xlsx_files(constants.XLSX_DIR)
    except Exception as e:
        logger.error(f"Failed to list .xlsx files: {e}")
        return

    num_threads = min(multiprocessing.cpu_count(), len(xlsx_files))

    with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = [executor.submit(process_file, file_path) for file_path in xlsx_files]
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logger.error(f"An error occurred during processing: {e}")

if __name__ == "__main__":
    main()
