#!/usr/bin/env python
# coding=utf-8

import os
import openpyxl
import gencommon
import concurrent.futures
from os import listdir
from multiprocessing import cpu_count
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Global Variables
KEY_ROW_IDX = 4
CPP_DIR = "generated/cpp/"
XLS_DIR = "xlsx/"


def get_column_names(sheet):
    """Get column names from the Excel sheet."""
    return [cell.value for cell in sheet[KEY_ROW_IDX] if cell.value is not None]


def get_key_row_data(row, column_names):
    """Extract key row data from the Excel sheet."""
    return {str(row[i].value): row[i].value for i in range(len(row)) if column_names[i].strip().lower() == "key"}


def get_workbook_data(workbook):
    """Extract data from the first sheet of the workbook."""
    workbook_data = {}
    sheet_names = workbook.sheetnames
    if sheet_names:  # Read only the first sheet
        sheet = workbook[sheet_names[0]]

        # Check if A5 cell value is 'multi' or None
        cell_value = sheet['A5'].value
        use_flat_multimap = cell_value is not None and cell_value.lower() == 'multi'
        first_19_rows_per_column = gencommon.get_first_19_rows_per_column(sheet)
        workbook_data[sheet_names[0]] = {
            'multi': use_flat_multimap,
            'get_first_19_rows_per_column': first_19_rows_per_column
        }
    return workbook_data


def get_cpp_type_name(type_name):
    if type_name == 'string':
        type_name = 'std::string'
    elif type_name.find('int') != -1:
        type_name = type_name + '_t'
    return type_name


def get_cpp_type_param_name_with_ref(type_name):
    if type_name == 'string':
        type_name = 'const std::string&'
    elif type_name.find('int') != -1:
        type_name = type_name + '_t'
    return type_name


def generate_cpp_header(datastring, sheetname, use_flat_multimap):
    """Generate C++ header file content."""
    sheet_name_lower = sheetname.lower()
    container_type = "unordered_multimap" if use_flat_multimap else "unordered_map"

    header_content = ["#pragma once", "#include <cstdint>", "#include <memory>", f'#include <unordered_map>',
                      f'#include "{sheet_name_lower}_config.pb.h"', f'class {sheetname}ConfigurationTable {{',
                      'public:', f'    using row_type = const {sheet_name_lower}_row*;',
                      f'    using kv_type = std::{container_type}<uint32_t, row_type>;',
                      f'    static {sheetname}ConfigurationTable& GetSingleton() {{ static {sheetname}ConfigurationTable singleton; return singleton; }}',
                      f'    const {sheet_name_lower}_table& All() const {{ return data_; }}',
                      f'    const std::pair<row_type, uint32_t> GetTable(uint32_t keyid);',
                      f'    const kv_type& KVData() const {{ return kv_data_; }}', '    void Load();', 'private:',
                      f'    {sheet_name_lower}_table data_;', '    kv_type kv_data_;\n\n', 'public:']

    for d in datastring:
        column_name = d[gencommon.COL_OBJ_COL_NAME]
        if d[gencommon.COL_OBJ_TABLE_KEY_INDEX] == gencommon.table_key:
            type_name = get_cpp_type_param_name_with_ref(d[gencommon.COL_OBJ_COL_TYPE])
            header_content.append(
                f'const std::pair<{sheetname}ConfigurationTable::row_type, uint32_t> GetBy{column_name.title()}({type_name} keyid) const;')
    header_content.append('\nprivate:')
    for d in datastring:
        column_name = d[gencommon.COL_OBJ_COL_NAME]
        if d[gencommon.COL_OBJ_TABLE_KEY_INDEX] == gencommon.table_key:
            type_name = get_cpp_type_name(d[gencommon.COL_OBJ_COL_TYPE])
            header_content.append(f'    std::{container_type}<{type_name}, row_type>  kv_{column_name}data_;')

    header_content.append('};')
    header_content.append(
        f'\ninline std::pair<{sheetname}ConfigurationTable::row_type, uint32_t> Get{sheetname}Table(uint32_t keyid) {{ return {sheetname}ConfigurationTable::GetSingleton().GetTable(keyid); }}'
    )
    header_content.append(
        f'\ninline const {sheet_name_lower}_table& Get{sheetname}AllTable() {{ return {sheetname}ConfigurationTable::GetSingleton().All(); }}'
    )

    return '\n'.join(header_content)


def generate_cpp_implementation(datastring, sheetname, use_flat_multimap):
    """Generate C++ implementation file content."""
    sheet_name_lower = sheetname.lower()
    container_type = "unordered_multimap" if use_flat_multimap else "unordered_map"

    cpp_content = [
        '#include "google/protobuf/util/json_util.h"',
        '#include "src/util/file2string.h"',
        '#include "muduo/base/Logging.h"',
        '#include "common_error_tip.pb.h"',
        f'#include "{sheet_name_lower}_config.h"\n',
        f'void {sheetname}ConfigurationTable::Load() {{',
        '    data_.Clear();',
        f'    const auto contents = File2String("config/generated/json/{sheet_name_lower}.json");',
        f'    if (const auto result = google::protobuf::util::JsonStringToMessage(contents.data(), &data_); '
        f'!result.ok()) {{',
        f'        LOG_FATAL << "{sheetname} " << result.message().data();',
        '    }',
        '    for (int32_t i = 0; i < data_.data_size(); ++i) {',
        '        const auto& row_data = data_.data(i);',
        '        kv_data_.emplace(row_data.id(), &row_data);\n\n',
    ]

    for d in datastring:
        column_name = d[gencommon.COL_OBJ_COL_NAME]
        if d[gencommon.COL_OBJ_TABLE_KEY_INDEX] == gencommon.table_key:
            cpp_content.append(f'        kv_{column_name}data_.emplace(row_data.{column_name}(), &row_data);')

    cpp_content.extend([
        '    }',
        '}\n\n',
        f'const std::pair<{sheetname}ConfigurationTable::row_type, uint32_t> {sheetname}ConfigurationTable::GetTable(uint32_t keyid) {{',
        '    const auto it = kv_data_.find(keyid);',
        '    if (it == kv_data_.end()) {',
        f'        LOG_ERROR << "{sheetname} table not found for ID: " << keyid;',
        '        return { nullptr, kInvalidTableId };',
        '    }',
        '    return { it->second, kOK };',
        '}\n\n',
    ])

    for d in datastring:
        column_name = d[gencommon.COL_OBJ_COL_NAME]
        if d[gencommon.COL_OBJ_TABLE_KEY_INDEX] == gencommon.table_key:
            type_name = get_cpp_type_param_name_with_ref(d[gencommon.COL_OBJ_COL_TYPE])
            cpp_content.extend([
                f'const std::pair<{sheetname}ConfigurationTable::row_type, uint32_t> '
                f'{sheetname}ConfigurationTable::GetBy{column_name.title()}({type_name} keyid) const{{',
                f'    const auto it = kv_{column_name}data_.find(keyid);',
                f'    if (it == kv_{column_name}data_.end()) {{',
                f'        LOG_ERROR << "{sheetname} table not found for ID: " << keyid;',
                '        return { nullptr, kInvalidTableId };',
                '    }',
                '    return { it->second, kOK };',
                '}\n',
            ])

    return '\n'.join(cpp_content)


def process_workbook(filename):
    """Process a single workbook file and generate corresponding header and implementation files."""
    try:
        workbook = openpyxl.load_workbook(filename)
    except Exception as e:
        logging.error(f"Failed to load workbook {filename}: {e}")
        return

    workbook_data = get_workbook_data(workbook)
    for sheetname, data in workbook_data.items():
        header_filename = f"{sheetname.lower()}_config.h"
        cpp_filename = f"{sheetname.lower()}_config.cpp"

        cpp_header_content = generate_cpp_header(data['get_first_19_rows_per_column'], sheetname, data['multi'])
        gencommon.mywrite(cpp_header_content, os.path.join(CPP_DIR, header_filename))

        cpp_implementation_content = generate_cpp_implementation(data['get_first_19_rows_per_column'], sheetname,
                                                                 data['multi'])
        gencommon.mywrite(cpp_implementation_content, os.path.join(CPP_DIR, cpp_filename))


def generate_all_config():
    """Generate header and implementation files for loading all configurations."""
    sheetnames = []
    dirfiles = listdir(XLS_DIR)
    files = sorted(dirfiles, key=lambda file: os.path.getsize(os.path.join(XLS_DIR, file)), reverse=True)

    for filename in files:
        filepath = os.path.join(XLS_DIR, filename)
        if filepath.endswith('.xlsx'):
            try:
                workbook = openpyxl.load_workbook(filepath)
                workbook_data = get_workbook_data(workbook)
                sheetnames.extend(workbook_data.keys())
            except Exception as e:
                logging.error(f"Failed to process file {filepath}: {e}")

    header_content = '#pragma once\n'
    header_content += 'void LoadAllConfig();\n'
    header_content += 'void LoadAllConfigAsyncWhenServerLaunch();\n'

    cpp_content = '#include "all_config.h"\n\n'
    cpp_content += '#include <thread>\n'
    cpp_content += '#include "muduo/base/CountDownLatch.h"\n\n'

    for item in sheetnames:
        cpp_content += f'#include "{item.lower()}_config.h"\n'

    cpp_content += 'void LoadAllConfig()\n{\n'
    for item in sheetnames:
        cpp_content += f'    {item}ConfigurationTable::GetSingleton().Load();\n'
    cpp_content += '}\n\n'

    cpucount = cpu_count()
    cpp_content += 'void LoadAllConfigAsyncWhenServerLaunch()\n{\n'
    cpp_content += f'    static muduo::CountDownLatch latch_({len(sheetnames)});\n'

    load_blocks = [[] for _ in range(cpucount)]
    for idx, item in enumerate(sheetnames):
        load_blocks[idx % cpucount].append(f'    {item}ConfigurationTable::GetSingleton().Load();\n')

    for block in load_blocks:
        if block:
            cpp_content += '\n    /// Begin\n'
            cpp_content += '    {\n'
            cpp_content += '        std::thread t([&]() {\n\n'
            cpp_content += ''.join(block)
            cpp_content += '            latch_.countDown();\n'
            cpp_content += '        });\n'
            cpp_content += '        t.detach();\n'
            cpp_content += '    }\n'
            cpp_content += '    /// End\n'

    cpp_content += '    latch_.wait();\n'
    cpp_content += '}\n'

    return header_content, cpp_content


def main():
    """Main function to generate all configuration files."""
    os.makedirs(CPP_DIR, exist_ok=True)

    # List of Excel files
    xlsx_files = [os.path.join(XLS_DIR, filename) for filename in listdir(XLS_DIR)
                  if filename.endswith('.xlsx')]

    # Process Excel files in parallel
    with concurrent.futures.ProcessPoolExecutor() as executor:
        executor.map(process_workbook, xlsx_files)

    # Generate header and implementation files for all configurations
    header_content, cpp_content = generate_all_config()
    gencommon.mywrite(header_content, os.path.join(CPP_DIR, "all_config.h"))
    gencommon.mywrite(cpp_content, os.path.join(CPP_DIR, "all_config.cpp"))


if __name__ == "__main__":
    main()
