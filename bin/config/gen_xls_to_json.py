#!/usr/bin/env python
# coding=utf-8

import os
import json
import logging
import concurrent.futures
from os import listdir
from os.path import isfile, join
from typing import Any, Union

import openpyxl
import generate_common  # Assuming generate_common contains the necessary functions
from common import constants
from config import XLSX_DIR

# Configuration Constants
JSON_DIR = "generated/json/"


# Setup Logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def get_column_names(sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    """
    Get column names from the sheet based on specified conditions.
    """
    return [
        sheet.cell(row=1, column=col_idx + 1).value
        if sheet.cell(row=4, column=col_idx + 1).value in constants.SERVER_GEN_TYPE
        else ""
        for col_idx in range(sheet.max_column)
    ]


def process_cell_value(cell: openpyxl.cell.cell.Cell, field_type: str) -> Union[float, int, str, Any]:
    """
    Process cell value to ensure the correct type based on the provided field_type.

    Args:
        cell (openpyxl.cell.cell.Cell): The Excel cell whose value needs processing.
        field_type (str): The type to which the cell value should be converted. Can be "float", "int", "string".

    Returns:
        Union[float, int, str, Any]: The processed cell value in the desired type.
    """
    cell_value = cell.value

    if field_type == "string":
        # 如果 field_type 是 "string"，并且 cell_value 为空（None 或 空字符串），返回 None
        if cell_value is None or cell_value == "":
            return None
        return str(cell_value)

    if field_type == "float" or field_type == "double":
        try:
            return float(cell_value)
        except ValueError:
            return None  # 如果无法转换为 float，返回 None 或进行适当的处理

    if isinstance(cell_value, float) and cell_value.is_integer():
        return int(cell_value)  # 如果值是浮动且是整数，转换为整数

    if isinstance(cell_value, int):
        return cell_value  # 如果值是整数，直接返回

    # 如果没有特定转换，返回原始的 cell_value
    return cell_value


def handle_map_field_data(cell, row_data, col_name, cell_value, map_field_data, column_names, prev_cell):
    """
    Handle data for map fields and update row data accordingly.
    """
    prev_column_name = column_names[prev_cell.col_idx - 1] if prev_cell else None
    if cell_value in (None, '') and cell.row >= generate_common.BEGIN_ROW_IDX:
        return

    prev_obj_name = generate_common.column_name_to_obj_name(prev_column_name, "_") if prev_column_name else None
    obj_name = generate_common.column_name_to_obj_name(col_name, "_")

    if generate_common.SET_CELL == map_field_data[col_name]:
        row_data.setdefault(col_name, {})[cell_value] = True
    elif (prev_column_name in map_field_data and generate_common.MAP_KEY_CELL == map_field_data[prev_column_name]
          and prev_obj_name == obj_name):
        row_data.setdefault(obj_name, {})[prev_cell.value] = cell_value


def handle_array_data(cell, row_data, col_name, cell_value):
    """
    Handle data for array fields and update row data accordingly.
    """
    if cell_value in (None, '', 0, -1) and cell.row >= generate_common.BEGIN_ROW_IDX:
        return
    row_data.setdefault(col_name, []).append(cell_value)


def handle_group_data(cell, row_data, col_name, cell_value, prev_cell):
    """
    Handle data for group fields and update row data accordingly.
    """
    if cell_value in (None, '', 0, -1) and cell.row >= generate_common.BEGIN_ROW_IDX:
        return

    obj_name = generate_common.column_name_to_obj_name(col_name, "_")
    member_dict = {col_name: cell_value}
    if obj_name in row_data:
        last_element = row_data[obj_name][-1]
        if col_name in last_element:
            row_data[obj_name].append(member_dict)
        else:
            last_element[col_name] = cell_value
    else:
        row_data[obj_name] = [member_dict]


def process_row(sheet, row, column_names):
    """
    Process an individual row to extract data based on column names.
    """
    sheet_data = generate_common.get_sheet_data(sheet, column_names)
    array_data = sheet_data[generate_common.SHEET_ARRAY_DATA_INDEX]
    field_type_data = sheet_data[generate_common.FILE_TYPE_INDEX]
    group_data = sheet_data[generate_common.SHEET_GROUP_ARRAY_DATA_INDEX]
    map_field_data = sheet_data[generate_common.MAP_TYPE_INDEX]

    row_data = {}
    prev_cell = None

    for counter, cell in enumerate(row):
        if counter >= len(column_names):
            logger.warning(f"Row {cell.row} in sheet '{sheet.title}' has more cells than column names.")
            break

        col_name = column_names[counter]
        if not col_name.strip():
            continue

        cell_value = process_cell_value(cell, field_type_data[col_name])
        if cell_value is None:
            continue
        if col_name in map_field_data or generate_common.is_key_in_map(group_data, col_name, map_field_data, column_names):
            handle_map_field_data(cell, row_data, col_name, cell_value, map_field_data, column_names, prev_cell)
        elif col_name in array_data:
            handle_array_data(cell, row_data, col_name, cell_value)
        elif generate_common.is_key_in_group_array(group_data, col_name, column_names):
            handle_group_data(cell, row_data, col_name, cell_value, prev_cell)
        else:
            if cell_value in (None, '') and cell.row >= generate_common.BEGIN_ROW_IDX:
                logger.error(f"Sheet '{sheet.title}', Cell {cell.coordinate} is empty or contains invalid data.")
            row_data[col_name] = cell_value

        prev_cell = cell

    return row_data


def extract_sheet_data(sheet, column_names):
    """
    Extract data from all rows in the sheet.
    """
    return [process_row(sheet, row, column_names) for row in
            sheet.iter_rows(min_row=generate_common.BEGIN_ROW_IDX + 1, values_only=False)]


def extract_workbook_data(workbook: openpyxl.Workbook) -> dict:
    """
    Extract data from the first sheet in the workbook.
    """
    workbook_data = {}
    if workbook.sheetnames:
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        if sheet.cell(row=1, column=1).value != "id":
            logger.error(f"{sheet_name} first column must be 'id'")
        else:
            column_names = get_column_names(sheet)
            workbook_data[sheet_name] = extract_sheet_data(sheet, column_names)
    else:
        logger.error("No sheets found in the workbook.")

    return workbook_data


def save_json(data: dict, file_path: str) -> None:
    """
    Save data to a JSON file with compact and readable formatting.
    """
    json_data = json.dumps({"data": data}, sort_keys=True, indent=1, separators=(',', ': '))
    json_data = json_data.replace('"[', '[').replace(']"', ']')  # Remove unnecessary quotes around lists
    json_data = json_data.replace('\r\n', '\n').replace('\r', '\n')  # Ensure only LF (\n) for newlines

    try:
        with open(file_path, 'w', encoding='utf-8', newline='') as f:
            f.write(json_data)
            logger.info(f"Generated JSON file: {file_path}")
    except IOError as e:
        logger.error(f"Error saving JSON file {file_path}: {e}")


def process_excel_file(file_path: str) -> None:
    """
    Process an individual Excel file and generate a JSON file for the first sheet.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        workbook_data = extract_workbook_data(workbook)

        for sheet_name, data in workbook_data.items():
            json_file_path = os.path.join(JSON_DIR, f"{sheet_name}.json")
            save_json(data, json_file_path)

    except Exception as e:
        logger.error(f"Failed to process file {file_path}: {e}")


def main() -> None:
    """
    Main function to process all Excel files in the specified directory.
    """
    os.makedirs(JSON_DIR, exist_ok=True)

    files = [join(XLSX_DIR, filename) for filename in listdir(XLSX_DIR) if
             isfile(join(XLSX_DIR, filename)) and filename.endswith('.xlsx')]

    num_threads = os.cpu_count() or 1
    with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
        executor.map(process_excel_file, files)


if __name__ == "__main__":
    main()
