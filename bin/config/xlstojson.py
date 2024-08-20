#!/usr/bin/env python
# coding=utf-8

import os
import json
import logging
import concurrent.futures
from os import listdir
from os.path import isfile, join
import openpyxl

import gencommon  # Assuming gencommon contains the necessary functions

# Configuration Constants
BEGIN_ROW_IDX = 9
JSON_DIR = "generated/json/"
XLSX_DIR = "xlsx/"
GEN_TYPE = "server"

# Setup Logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def get_column_names(sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    """
    Get column names from the sheet based on specified conditions.
    """
    return [
        sheet.cell(row=1, column=col_idx + 1).value
        if sheet.cell(row=4, column=col_idx + 1).value in ["common", GEN_TYPE]
        else ""
        for col_idx in range(sheet.max_column)
    ]


def process_cell_value(cell: openpyxl.cell.cell.Cell) -> int | str | None:
    """
    Process cell value to ensure correct type.
    """
    cell_value = cell.value
    if isinstance(cell_value, float) and cell_value.is_integer():
        return int(cell_value)
    return cell_value


def handle_map_field_data(cell, row_data, col_name, cell_value, map_field_data, column_names, prev_cell):
    """
    Handle data for map fields and update row data accordingly.
    """
    prev_column_name = column_names[prev_cell.col_idx - 1] if prev_cell else None

    if cell_value in (None, '') and cell.row >= BEGIN_ROW_IDX:
        return

    prev_obj_name = gencommon.column_name_to_obj_name(prev_column_name, "_") if prev_column_name else None
    obj_name = gencommon.column_name_to_obj_name(col_name, "_")

    if gencommon.set_flag == map_field_data[col_name]:
        row_data.setdefault(col_name, {})[cell_value] = True
    elif (prev_column_name in map_field_data and gencommon.map_key_flag == map_field_data[prev_column_name] and
          prev_obj_name == obj_name):
        row_data.setdefault(obj_name, {})[prev_cell.value] = cell_value


def handle_array_data(cell, row_data, col_name, cell_value):
    """
    Handle data for array fields and update row data accordingly.
    """
    if cell_value in (None, '') and cell.row >= BEGIN_ROW_IDX:
        return
    if cell_value in (0, -1):
        return

    row_data.setdefault(col_name, []).append(cell_value)


def handle_group_data(cell, row_data, col_name, cell_value, prev_cell):
    """
    Handle data for group fields and update row data accordingly.
    """
    if cell_value in (None, '') and cell.row >= BEGIN_ROW_IDX:
        return
    if cell_value in (0, -1):
        return

    obj_name = gencommon.column_name_to_obj_name(col_name, "_")
    member_dict = {col_name: cell_value}
    if obj_name in row_data:
        last_element = row_data[obj_name][-1]
        if col_name in last_element:
            row_data[obj_name].append(member_dict)
        else:
            last_element[col_name] = cell_value
    else:
        row_data[obj_name] = [member_dict]


def process_row(sheet, row, column_names):
    """
    Process an individual row to extract data based on column names.
    """
    sheet_data = gencommon.get_sheet_data(sheet, column_names)
    array_data = sheet_data[gencommon.sheet_array_data_index]
    group_data = sheet_data[gencommon.sheet_group_array_data_index]
    map_field_data = sheet_data[gencommon.map_type_index]

    row_data = {}
    prev_cell = None

    for counter, cell in enumerate(row):
        if counter >= len(column_names):
            logger.warning(f"Row {cell.row} in sheet '{sheet.title}' has more cells than column names.")
            break

        col_name = column_names[counter]
        if col_name == 'designer' or not col_name.strip():
            continue

        cell_value = process_cell_value(cell)
        if col_name in map_field_data or gencommon.is_key_in_map(group_data, col_name, map_field_data, column_names):
            handle_map_field_data(cell, row_data, col_name, cell_value, map_field_data, column_names, prev_cell)
        elif col_name in array_data:
            handle_array_data(cell, row_data, col_name, cell_value)
        elif gencommon.is_key_in_group_array(group_data, col_name, column_names):
            handle_group_data(cell, row_data, col_name, cell_value, prev_cell)
        else:
            if cell_value in (None, '') and cell.row >= BEGIN_ROW_IDX:
                logger.error(f"Sheet '{sheet.title}', Cell {cell.coordinate} is empty or contains invalid data.")
            row_data[col_name] = cell_value

        prev_cell = cell

    return row_data


def extract_sheet_data(sheet, column_names):
    """
    Extract data from all rows in the sheet.
    """
    return [process_row(sheet, row, column_names) for row in
            sheet.iter_rows(min_row=BEGIN_ROW_IDX + 1, values_only=False)]


def extract_workbook_data(workbook: openpyxl.Workbook) -> dict:
    """
    Extract data from all sheets in the workbook.
    """
    workbook_data = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet.cell(row=1, column=1).value != "id":
            logger.error(f"{sheet_name} first column must be 'id'")
            continue

        column_names = get_column_names(sheet)
        workbook_data[sheet_name] = extract_sheet_data(sheet, column_names)

    return workbook_data


def save_json(data: dict, file_path: str) -> None:
    """
    Save data to a JSON file with compact and readable formatting.
    """
    json_data = json.dumps({"data": data}, sort_keys=True, indent=1, separators=(',', ': '))
    json_data = json_data.replace('"[', '[').replace(']"', ']')  # Remove unnecessary quotes around lists
    json_data = json_data.replace('\r\n', '\n').replace('\r', '\n')  # Ensure only LF (\n) for newlines

    try:
        with open(file_path, 'w', encoding='utf-8', newline='') as f:
            f.write(json_data)
            logger.info(f"Generated JSON file: {file_path}")
    except IOError as e:
        logger.error(f"Error saving JSON file {file_path}: {e}")


def process_excel_file(file_path: str) -> None:
    """
    Process an individual Excel file and generate JSON files for each sheet.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        workbook_data = extract_workbook_data(workbook)

        for sheet_name, data in workbook_data.items():
            json_file_path = os.path.join(JSON_DIR, f"{sheet_name}.json")
            save_json(data, json_file_path)

    except Exception as e:
        logger.error(f"Failed to process file {file_path}: {e}")


def main() -> None:
    """
    Main function to process all Excel files in the specified directory.
    """
    os.makedirs(JSON_DIR, exist_ok=True)

    files = [join(XLSX_DIR, filename) for filename in listdir(XLSX_DIR) if
             isfile(join(XLSX_DIR, filename)) and filename.endswith('.xlsx')]

    num_threads = os.cpu_count() or 1
    with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
        executor.map(process_excel_file, files)


if __name__ == "__main__":
    main()
