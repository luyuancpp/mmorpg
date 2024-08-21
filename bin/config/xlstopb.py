#!/usr/bin/env python
# coding=utf-8

import os
import openpyxl
import logging
import concurrent.futures
import multiprocessing
from typing import Dict, List, Optional
from os import listdir
from os.path import isfile, join
import gencommon

# Configuration Constants
PROTO_DIR = "generated/proto/"
XLSX_DIR = "xlsx/"

# Setup Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Indices for data extraction


def get_workbook_data(workbook: openpyxl.Workbook) -> Dict[str, Dict]:
    """Extract data from all sheets in the workbook."""
    data = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if validate_sheet(sheet):
            column_names = gencommon.get_column_names(sheet)
            if column_names:
                sheet_data = gencommon.get_sheet_data(sheet, column_names)
                data[sheet_name] = sheet_data
    return data

def validate_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    """Ensure the sheet's first column header is 'id'."""
    if sheet.cell(row=1, column=1).value != "id":
        logger.error(f"Sheet '{sheet.title}' first column must be 'id'")
        return False
    return True

def generate_proto_file(data: Dict, sheet_name: str) -> Optional[str]:
    """Generate .proto file content based on sheet data."""
    try:
        proto_content = create_proto_header()
        column_names = data[gencommon.SHEET_COLUM_NAME_INDEX]

        proto_content += generate_group_messages(data, column_names)
        proto_content += generate_row_message(sheet_name, data, column_names)
        proto_content += generate_table_message(sheet_name)

        return proto_content
    except Exception as e:
        logger.error(f"Error generating proto content for '{sheet_name}': {e}")
        return None

def create_proto_header() -> str:
    """Create the initial header for the .proto file."""
    return (
        'syntax = "proto3";\n\n'
        'option go_package = "pb/game";\n\n'
    )

def generate_group_messages(data: Dict, column_names: List[str]) -> str:
    """Generate messages for grouped data."""
    proto_content = ''
    group_data = data[gencommon.SHEET_GROUP_ARRAY_DATA_INDEX]
    for k, v in group_data.items():
        obj_name = gencommon.set_to_string(
            gencommon.find_common_words(column_names[v[0]], column_names[v[1]], '_')
        )
        proto_content += f'message {obj_name} {{\n'
        for i, column in enumerate(v):
            name = column_names[column]
            proto_content += f'\t{data[0][name]} {name} = {i + 1};\n'
        proto_content += '}\n\n'
    return proto_content

def generate_row_message(sheet_name: str, data: Dict, column_names: List[str]) -> str:
    """Generate the row message for the .proto file."""
    proto_content = f'message {sheet_name.lower()}_row {{\n'
    field_index = 1

    for key, _ in data[0].items():
        if not is_excluded_owner(data[gencommon.OWNER_INDEX], key):
            field_content = format_field(data, key, column_names, field_index)
            if field_content:
                proto_content += field_content
                field_index += 1

    proto_content += '}\n\n'
    return proto_content

def is_excluded_owner(owner_data: Dict, key: str) -> bool:
    """Check if the key is excluded based on owner data."""
    return owner_data.get(key, '').strip() in ('client', 'design')

def format_field(data: Dict, key: str, column_names: List[str], field_index: int) -> str:
    """Format a field for the .proto file based on its type."""
    if key in data[gencommon.MAP_TYPE_INDEX]:
        return format_map_field(data, key, column_names, field_index)
    elif key in data[gencommon.SHEET_ARRAY_DATA_INDEX]:
        return f'\trepeated {data[0][key]} {key} = {field_index};\n'
    elif gencommon.is_key_in_group_array(data[gencommon.SHEET_GROUP_ARRAY_DATA_INDEX], key, column_names):
        return format_group_array_field(data, key, column_names, field_index)
    else:
        return f'\t{data[0][key]} {key} = {field_index};\n'

def format_map_field(data: Dict, key: str, column_names: List[str], field_index: int) -> str:
    """Format a map field for the .proto file."""
    map_type = data[gencommon.MAP_TYPE_INDEX][key]
    if map_type == gencommon.set_flag:
        return f'\tmap <{data[0][key]}, bool> {key} = {field_index};\n'
    elif map_type == gencommon.map_key_flag:
        value_type = data[gencommon.FILE_TYPE_INDEX]
        if key in data[gencommon.SHEET_GROUP_ARRAY_DATA_INDEX]:
            value = data[gencommon.SHEET_GROUP_ARRAY_DATA_INDEX][key]
            key_name = column_names[value[0]]
            value_name = column_names[value[1]]
            obj_name = gencommon.column_name_to_obj_name(key_name, '_')
            return f'\tmap <{value_type[key_name]}, {value_type[value_name]}> {obj_name} = {field_index};\n'
    return ''

def format_group_array_field(data: Dict, key: str, column_names: List[str], field_index: int) -> str:
    """Format a group array field for the .proto file."""
    if key not in data[gencommon.SHEET_GROUP_ARRAY_DATA_INDEX]:
        return ''
    value = data[gencommon.SHEET_GROUP_ARRAY_DATA_INDEX][key]
    key_name = column_names[value[0]]
    obj_name = gencommon.column_name_to_obj_name(key_name, '_')
    return f'\trepeated {obj_name} {obj_name} = {field_index};\n'

def generate_table_message(sheet_name: str) -> str:
    """Generate the table message for the .proto file."""
    return (
        f'message {sheet_name}_table {{\n'
        f'\trepeated {sheet_name.lower()}_row data = 1;\n'
        '}\n'
    )

def process_file(file_path: str) -> None:
    """Process an individual Excel file to generate .proto files."""
    try:
        workbook = openpyxl.load_workbook(file_path)
        workbook_data = get_workbook_data(workbook)

        for sheet_name, data in workbook_data.items():
            proto_content = generate_proto_file(data, sheet_name.lower())
            if proto_content:
                save_proto_file(proto_content, sheet_name.lower())
    except Exception as e:
        logger.error(f"Error processing file '{file_path}': {e}")

def save_proto_file(content: str, sheet_name: str) -> None:
    """Save the generated .proto content to a file."""
    proto_file_path = os.path.join(PROTO_DIR, f'{sheet_name}_config.proto')
    try:
        with open(proto_file_path, 'w', encoding='utf-8') as proto_file:
            proto_file.write(content)
            logger.info(f"Generated .proto file: {proto_file_path}")
    except Exception as e:
        logger.error(f"Error saving proto file '{proto_file_path}': {e}")

def get_xlsx_files(directory: str) -> List[str]:
    """List all .xlsx files in the specified directory."""
    return [join(directory, filename) for filename in listdir(directory) if isfile(join(directory, filename)) and filename.endswith('.xlsx')]

def main() -> None:
    """Main function to process all Excel files and generate .proto files."""
    try:
        os.makedirs(PROTO_DIR, exist_ok=True)
    except Exception as e:
        logger.error(f"Failed to create proto directory: {e}")
        return

    try:
        xlsx_files = get_xlsx_files(XLSX_DIR)
    except Exception as e:
        logger.error(f"Failed to list .xlsx files: {e}")
        return

    num_threads = min(multiprocessing.cpu_count(), len(xlsx_files))
    with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = [executor.submit(process_file, file_path) for file_path in xlsx_files]
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logger.error(f"An error occurred during processing: {e}")

if __name__ == "__main__":
    main()
