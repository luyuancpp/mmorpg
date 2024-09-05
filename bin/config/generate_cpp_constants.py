#!/usr/bin/env python
# coding=utf-8

import os
import json
import logging
import concurrent.futures
from os import listdir
from os.path import isfile, join
import openpyxl
import multiprocessing
from typing import List, Optional

import gen_common  # Assuming gen_common contains the necessary functions
from common import constants

# Setup Logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelToCppConverter:
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.workbook = openpyxl.load_workbook(excel_file)
        self.sheet = self.workbook.sheetnames[0]
        self.worksheet = self.workbook[self.sheet]

    def generate_cpp_constants(self) -> str:
        cpp_constants = "#pragma once\n\n"
        headers = [cell.value for cell in self.worksheet[1]]
        constants_name_index = headers.index('constants_name') if 'constants_name' in headers else None

        for row in self.worksheet.iter_rows(min_row=20, values_only=True):
            id_value = row[0]
            if id_value is None:
                continue  # Skip rows with no ID

            constant_name = self._generate_constant_name(row, constants_name_index, id_value)
            cpp_constant = f"constexpr uint32_t {constant_name} = {id_value};\n"
            cpp_constants += cpp_constant

        return cpp_constants

    def _generate_constant_name(self, row: tuple, constants_name_index: Optional[int], id_value: int) -> str:
        if constants_name_index is not None and row[constants_name_index]:
            return f'k{self.sheet}_{row[constants_name_index]}'
        return f"k{self.sheet}_{id_value}"

    def save_cpp_constants_to_file(self, cpp_constants: str) -> None:
        output_file = constants.GENERATOR_CONSTANTS_NAME_DIR + self.sheet.lower() + '_table_id_constants.h'
        with open(output_file, 'w') as file:
            file.write(cpp_constants)


def get_xlsx_files(directory: str) -> List[str]:
    """List all .xlsx files in the specified directory."""
    return [join(directory, filename) for filename in listdir(directory)
            if isfile(join(directory, filename)) and filename.endswith('.xlsx')]


def process_file(excel_file: str) -> None:
    converter = ExcelToCppConverter(excel_file)
    cpp_constants = converter.generate_cpp_constants()
    converter.save_cpp_constants_to_file(cpp_constants)


def main() -> None:
    os.makedirs(constants.GENERATOR_CONSTANTS_NAME_DIR, exist_ok=True)

    try:
        xlsx_files = get_xlsx_files(constants.XLSX_DIR)
    except Exception as e:
        logger.error(f"Failed to list .xlsx files: {e}")
        return

    num_threads = min(multiprocessing.cpu_count(), len(xlsx_files))

    with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = [executor.submit(process_file, file_path) for file_path in xlsx_files]
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logger.error(f"An error occurred during processing: {e}")


if __name__ == "__main__":
    main()
