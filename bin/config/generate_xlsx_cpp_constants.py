#!/usr/bin/env python
# coding=utf-8

import os
import logging
import openpyxl
from typing import Optional
from jinja2 import Environment, FileSystemLoader
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

import utils
import generate_common
from common import constants

# Set up logging
logging.basicConfig(
    level=logging.DEBUG,  # Increase log level for debugging
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ExcelToCppConverter:
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.is_global_file = 'globalvariable' in os.path.basename(excel_file).lower()
        try:
            self.workbook = openpyxl.load_workbook(excel_file)
            self.sheet = self.workbook.sheetnames[0]
            self.worksheet = self.workbook[self.sheet]
            self.constants_name_index = self._find_constants_name_index()
        except Exception as e:
            logger.error(f"Failed to load Excel file: {excel_file}, error: {e}")
            raise

    def _find_constants_name_index(self) -> Optional[int]:
        try:
            if not self.worksheet or self.worksheet.max_row < 1:
                logger.warning(f"Worksheet '{self.sheet}' is empty or has no data")
                return None

            first_row = self.worksheet[1]
            headers = [str(cell.value).strip() if cell.value else '' for cell in first_row]
            return headers.index('constants_name') if 'constants_name' in headers else None
        except Exception as e:
            logger.error(f"Error finding 'constants_name' column: {e}")
            return None

    def should_process(self) -> bool:
        return self.constants_name_index is not None

    def generate_cpp_constants(self) -> str:
        constants_list = []
        for row in self.worksheet.iter_rows(min_row=20, values_only=True):
            id_value = row[0]
            if id_value is None:
                continue
            constant_name = self._generate_constant_name(row, id_value)
            constants_list.append({'name': constant_name, 'value': id_value})

        # Explicitly specify UTF-8 encoding for templates to avoid decoding errors
        env = Environment(loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'))
        template = env.get_template("constants.h.j2")
        return template.render(constants=constants_list)

    def _generate_constant_name(self, row: tuple, id_value: int) -> str:
        if self.constants_name_index is not None and row[self.constants_name_index]:
            return f'k{self.sheet}_{row[self.constants_name_index]}'
        return f'k{self.sheet}_{id_value}'

    def save_cpp_constants_to_file(self, cpp_code: str) -> None:
        output_path = os.path.join(
            constants.GENERATOR_CONSTANTS_NAME_DIR,
            f"{self.sheet.lower()}_table_id_constants.h"
        )
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(cpp_code)

    def close(self):
        if hasattr(self, 'workbook'):
            self.workbook.close()

    def generate_and_save_cpp_constants(self):
        if not self.should_process():
            logger.info(f"No 'constants_name' column found, skipping file: {self.excel_file}")
            return

        if self.is_global_file:
            # 按 constants_name 拆分生成多个文件
            constants_map = {}  # key: (first_word, second_word), value: list of constants
            for row in self.worksheet.iter_rows(min_row=20, values_only=True):
                id_value = row[0]
                if id_value is None:
                    continue
                if not row[self.constants_name_index]:
                    continue
                const_name_full = str(row[self.constants_name_index])
                parts = const_name_full.split('_')
                if len(parts) < 2:
                    logger.warning(f"Invalid constants_name format (need at least two parts): {const_name_full}")
                    continue
                key = (parts[0], parts[1])
                constant_name = f'k{self.sheet}_{const_name_full}'
                if key not in constants_map:
                    constants_map[key] = []
                constants_map[key].append({'name': constant_name, 'value': id_value})

            env = Environment(loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'))
            template = env.get_template("constants.h.j2")

            for key, const_list in constants_map.items():
                first_word, second_word = key
                cpp_code = template.render(constants=const_list)
                filename = f"global_{first_word}_{second_word}.h".lower()  # 👈 转小写
                output_path = os.path.join(constants.GENERATOR_CONSTANTS_NAME_DIR, filename)
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(cpp_code)
                logger.info(f"Generated file: {output_path}")

        else:
            # 普通文件处理逻辑
            constants_list = []
            for row in self.worksheet.iter_rows(min_row=20, values_only=True):
                id_value = row[0]
                if id_value is None:
                    continue
                constant_name = self._generate_constant_name(row, id_value)
                constants_list.append({'name': constant_name, 'value': id_value})

            env = Environment(loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'))
            template = env.get_template("constants.h.j2")
            cpp_code = template.render(constants=constants_list)

            output_path = os.path.join(
                constants.GENERATOR_CONSTANTS_NAME_DIR,
                f"{self.sheet.lower()}_table_id_constants.h"
            )
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(cpp_code)
            logger.info(f"Generated file: {output_path}")

    # 处理函数改成调用这个新方法


def process_file(file_path: str):
    logger.info(f"Start processing file: {file_path}")

    if not os.path.isfile(file_path):
        logger.error(f"File not found: {file_path}")
        return

    try:
        converter = ExcelToCppConverter(file_path)
        converter.generate_and_save_cpp_constants()
        logger.info(f"Finished processing: {file_path}")
    except Exception as e:
        logger.error(f"Error processing file: {file_path}, error: {e}")
    finally:
        if 'converter' in locals():
            converter.close()

def main():
    try:
        os.makedirs(constants.GENERATOR_CONSTANTS_NAME_DIR, exist_ok=True)

        xlsx_files = utils.get_xlsx_files(constants.XLSX_DIR)
        if not xlsx_files:
            logger.warning("No Excel files found to process")
            return

        for file in xlsx_files:
            process_file(file)

    except Exception as e:
        logger.error(f"Error during program execution: {e}")
    finally:
        logging.shutdown()


if __name__ == "__main__":
    main()
