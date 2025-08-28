#!/usr/bin/env python
# coding=utf-8

import concurrent.futures
import os
import openpyxl
import generate_common
import concurrent.futures
from pathlib import Path
from multiprocessing import cpu_count
from pathlib import Path
from multiprocessing import cpu_count
import openpyxl
import logging
from jinja2 import Environment, FileSystemLoader


# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Global Variables
KEY_ROW_IDX = 4
CPP_DIR = Path("generated/cpp")
XLS_DIR = Path("xlsx")

def get_column_names(sheet):
    """Get column names from the Excel sheet."""
    return [cell.value for cell in sheet[KEY_ROW_IDX] if cell.value is not None]


def get_key_row_data(row, column_names):
    """Extract key row data from the Excel sheet."""
    return {str(row[i].value): row[i].value for i in range(len(row)) if column_names[i].strip().lower() == "key"}


def get_workbook_data(workbook):
    """Extract data from the first sheet of the workbook."""
    workbook_data = {}
    sheet_names = workbook.sheetnames
    if sheet_names:  # Read only the first sheet
        sheet = workbook[sheet_names[0]]

        # Check if A5 cell value is 'multi' or None
        cell_value = sheet['A5'].value
        use_flat_multimap = cell_value is not None and cell_value.lower() == 'multi'
        first_19_rows_per_column = generate_common.get_first_19_rows_per_column(sheet)
        workbook_data[sheet_names[0]] = {
            'multi': use_flat_multimap,
            'get_first_19_rows_per_column': first_19_rows_per_column
        }
    return workbook_data


def get_cpp_type_name(type_name):
    """Return C++ type name based on the provided type_name."""
    if type_name == 'string':
        return 'std::string'
    elif 'int' in type_name:
        return f'{type_name}_t'
    return type_name


def get_cpp_type_param_name_with_ref(type_name):
    """Return C++ type parameter name with reference."""
    if type_name == 'string':
        return 'const std::string&'
    elif 'int' in type_name:
        return f'{type_name}_t'
    return type_name

def process_workbook(filename):
    """Process a single workbook file and generate corresponding header and implementation files."""
    try:
        workbook = openpyxl.load_workbook(filename)
        workbook_data = get_workbook_data(workbook)
        for sheetname, data in workbook_data.items():
            header_filename = f"{sheetname.lower()}_table.h"
            cpp_filename = f"{sheetname.lower()}_table.cpp"

            # Create a Jinja2 environment and load the templates
            env = Environment(loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'), auto_reload=True)

            # Render the header templates
            header_template = env.get_template('config_template.h.jinja')
            header_content = header_template.render(
                datastring=data['get_first_19_rows_per_column'],
                sheetname=sheetname,
                use_flat_multimap=data['multi'],
                generate_common=generate_common,
                get_cpp_type_param_name_with_ref=get_cpp_type_param_name_with_ref,
                get_cpp_type_name=get_cpp_type_name
            )

            generate_common.mywrite(header_content, CPP_DIR / header_filename)

            # Render the implementation templates
            implementation_template = env.get_template('config_template.cpp.jinja')
            implementation_content = implementation_template.render(
                datastring=data['get_first_19_rows_per_column'],
                sheetname=sheetname,
                generate_common=generate_common,
                get_cpp_type_param_name_with_ref=get_cpp_type_param_name_with_ref,
                get_cpp_type_name=get_cpp_type_name

            )

            generate_common.mywrite(implementation_content, CPP_DIR / cpp_filename)
    except Exception as e:
        logging.error(f"Failed to load or process workbook {filename}: {e}")


def generate_all_config():
    # 收集所有 sheet 名
    sheetnames = set()
    xlsx_files = sorted(XLS_DIR.glob('*.xlsx'), key=lambda f: f.stat().st_size, reverse=True)
    for filepath in xlsx_files:
        try:
            workbook = openpyxl.load_workbook(filepath)
            workbook_data = get_workbook_data(workbook)
            sheetnames.update(workbook_data.keys())
        except Exception as e:
            logging.error(f"Failed to process file {filepath}: {e}")

    sheetnames = sorted(sheetnames)  # 保持稳定输出
    cpucount = cpu_count()

    # 初始化模板引擎
    env = Environment(loader=FileSystemLoader(generate_common.TEMPLATE_DIR, encoding='utf-8'))
    header_template = env.get_template("all_config.h.jinja")
    cpp_template = env.get_template("all_config.cpp.jinja")

    header_content = header_template.render()
    cpp_content = cpp_template.render(sheetnames=sheetnames, cpucount=cpucount)

    return header_content, cpp_content

def main():
    """Main function to generate all configuration files."""
    CPP_DIR.mkdir(parents=True, exist_ok=True)

    # List of Excel files
    xlsx_files = [XLS_DIR / filename for filename in os.listdir(XLS_DIR) if filename.endswith('.xlsx')]

    # Process Excel files in parallel
    with concurrent.futures.ProcessPoolExecutor() as executor:
        executor.map(process_workbook, xlsx_files)




if __name__ == "__main__":
    main()
