#!/usr/bin/env python
# coding=utf-8

import os
import openpyxl
import gencommon  # Ensure gencommon provides BEGIN_ROW_IDX and mywrite functions
import logging
from os import listdir
from os.path import isfile, join

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

cpp_dir = "generated/cpp/"
xlsx_dir = "xlsx/"
gen_file_list = ["global_variable"]

def generate_id_enum(sheet):
    """Generate a C++ enum definition based on the provided sheet."""
    n_rows = sheet.max_row
    name = sheet.title
    file_str = "#pragma once\n"
    file_str += f"enum e_{name}_configid : uint32_t\n{{\n"

    for idx in range(gencommon.BEGIN_ROW_IDX, n_rows + 1):
        cell_value = sheet.cell(row=idx, column=1).value
        if cell_value is not None:
            file_str += f"    {name}_config_id_{int(cell_value)},\n"

    file_str += "};\n"
    return file_str

def generate_id_cpp(workbook):
    """Generate C++ enum files for the first sheet in the workbook."""
    workbook_data = {}
    first_sheet_name = workbook.sheetnames[0]  # Only get the first sheet
    if first_sheet_name in gen_file_list:
        sheet = workbook[first_sheet_name]
        file_data = generate_id_enum(sheet)
        workbook_data[first_sheet_name] = file_data
    return workbook_data

def main():
    """Main function to process all Excel files and generate C++ enum files."""
    if not os.path.exists(cpp_dir):
        os.makedirs(cpp_dir)

    for filename in listdir(xlsx_dir):
        full_path = os.path.join(xlsx_dir, filename)
        if isfile(full_path) and filename.endswith('.xlsx'):
            try:
                workbook = openpyxl.load_workbook(full_path)
                workbook_data = generate_id_cpp(workbook)
                for sheet_name, data in workbook_data.items():
                    output_file_path = os.path.join(cpp_dir, f"{sheet_name}_config_id.h")
                    gencommon.mywrite(data, output_file_path)
                    logger.info(f"Generated C++ enum file: {output_file_path}")
            except Exception as e:
                logger.error(f"Failed to process file {full_path}: {e}")

if __name__ == "__main__":
    main()
