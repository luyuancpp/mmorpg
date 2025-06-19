#!/usr/bin/env python
# coding=utf-8

import os
import json
import logging
import concurrent.futures
from os import listdir
from os.path import isfile, join
import openpyxl
import multiprocessing
from typing import List, Optional, Dict

import generate_common  # Assuming generate_common contains the necessary functions
from common import constants
from jinja2 import Environment, FileSystemLoader

# Setup Logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)



class ExcelToCppConverter:
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.workbook = openpyxl.load_workbook(excel_file)
        self.sheet = self.workbook.sheetnames[0]
        self.worksheet = self.workbook[self.sheet]
        self.bit_index_col = self._find_bit_index_column()
        self.mapping_file = join(constants.GENERATOR_TABLE_INDEX_MAPPING_DIR, f"{self.sheet.lower()}_mapping.json")

        # Initialize Jinja2 environment
        self.template_env = Environment(
            loader=FileSystemLoader(generate_common.TEMPLATE_DIR),  # Path to your template folder
            autoescape=True
        )

    def _find_bit_index_column(self) -> Optional[int]:
        """Find the index of the column where the 7th row contains 'bit_index'."""
        headers = [cell.value for cell in self.worksheet[1]]
        for col_idx in range(len(headers)):
            cell_value = self.worksheet.cell(row=generate_common.XLSX_TABLE_BIT_BEGIN_INDEX, column=col_idx + 1).value
            if cell_value is not None and cell_value.strip().lower() == 'bit_index':
                return col_idx
        return None

    def _load_existing_mapping(self) -> Dict[int, int]:
        """Load existing ID to index mapping from a JSON file."""
        if os.path.exists(self.mapping_file):
            with open(self.mapping_file, 'r') as file:
                try:
                    # Load the JSON data
                    data = json.load(file)
                    return {int(k): v for k, v in data.items()}  # Convert keys to integers
                except json.JSONDecodeError:
                    logger.error("Error: JSON file is not valid.")
                    return {}
        return {}

    def _save_mapping(self, mapping: Dict[int, int]) -> None:
        """Save the ID to index mapping to a JSON file."""
        with open(self.mapping_file, 'w') as file:
            json.dump(mapping, file, indent=4)

    def _find_unused_indexes(self, id_to_index: Dict[int, int]) -> List[int]:
        """Find and return a list of unused indexes."""
        used_indexes = set(id_to_index.values())
        all_indexes = set(range(len(id_to_index)))
        unused_indexes = sorted(all_indexes - used_indexes)
        return unused_indexes

    def _find_max_bit_index(self) -> int:
        """Find the maximum bit index in the 7th row."""
        max_bit_index = -1
        if self.bit_index_col is not None:
            for row in self.worksheet.iter_rows(min_row=20, values_only=True):
                bit_index = row[self.bit_index_col]
                if isinstance(bit_index, int) and bit_index > max_bit_index:
                    max_bit_index = bit_index
        return max_bit_index

    def should_process(self) -> bool:
        """Check if the worksheet contains a valid 'bit_index' column."""
        return self.bit_index_col is not None

    def generate_cpp_constants(self) -> str:
        """Generate C++ constants from the Excel data using Jinja2 template."""
        id_to_index = self._load_existing_mapping()
        unused_indexes = self._find_unused_indexes(id_to_index)
        current_index = max(id_to_index.values(), default=-1) + 1

        # Iterate through rows to assign ID to index
        for row in self.worksheet.iter_rows(min_row=20, values_only=True):
            id_value = row[0]
            if id_value is None:
                continue  # Skip rows with no ID

            if id_value not in id_to_index:
                if unused_indexes:
                    index = unused_indexes.pop(0)
                else:
                    index = current_index
                    current_index += 1
                id_to_index[id_value] = index

        # Render the template
        template = self.template_env.get_template("cpp_config_id_bit_template.h.j2")
        cpp_constants = template.render(
            sheet=self.sheet,
            id_to_index=id_to_index,
            max_bit_index=self._find_max_bit_index()
        )

        # Save the mapping
        self._save_mapping(id_to_index)
        return cpp_constants

    def save_cpp_constants_to_file(self, cpp_constants: str) -> None:
        """Save the generated C++ constants to a file."""
        output_file = join(constants.GENERATOR_TABLE_INDEX_DIR, f"{self.sheet.lower()}_table_id_bit_index.h")
        with open(output_file, 'w') as file:
            file.write(cpp_constants)


def get_xlsx_files(directory: str) -> List[str]:
    """List all .xlsx files in the specified directory."""
    return [join(directory, filename) for filename in listdir(directory)
            if isfile(join(directory, filename)) and filename.endswith('.xlsx')]

def process_file(excel_file: str) -> None:
    """Process each Excel file and generate the corresponding C++ constants file."""
    converter = ExcelToCppConverter(excel_file)
    if converter.should_process():
        cpp_constants = converter.generate_cpp_constants()
        converter.save_cpp_constants_to_file(cpp_constants)
    else:
        logger.debug(f"Skipping file {excel_file} as it does not contain a valid 'bit_index' value in the 7th row.")


def main() -> None:
    """Main function to process all Excel files."""
    os.makedirs(constants.GENERATOR_TABLE_INDEX_DIR, exist_ok=True)
    os.makedirs(constants.GENERATOR_TABLE_INDEX_MAPPING_DIR, exist_ok=True)

    try:
        xlsx_files = get_xlsx_files(constants.XLSX_DIR)
        num_threads = min(multiprocessing.cpu_count(), len(xlsx_files))

        with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
            try:
                futures = [executor.submit(process_file, file_path) for file_path in xlsx_files]
                for future in concurrent.futures.as_completed(futures):
                    try:
                        future.result()
                    except Exception as e:
                        logger.error(f"任务执行失败: {str(e)}")
            except Exception as e:
                logger.error(f"线程池执行失败: {str(e)}")
            finally:
                # 确保所有任务都完成
                executor.shutdown(wait=True)
    except Exception as e:
        logger.error(f"主程序执行失败: {str(e)}")


if __name__ == "__main__":
    main()
