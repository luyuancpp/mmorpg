#!/usr/bin/env python
# coding=utf-8

import os
import openpyxl
import logging
from os import listdir
from os.path import isfile, join

import generate_common  # 你项目已有模块，需包含 BEGIN_ROW_IDX 和 mywrite
from config import XLSX_DIR

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

GO_DIR = "generated/go/table_id/"

def to_pascal_case(s: str) -> str:
    """转为 PascalCase，例如 global_variable -> GlobalVariable"""
    return ''.join(word.capitalize() for word in s.split('_'))

def generate_go_const(sheet):
    """根据 Excel sheet 生成 Go 常量定义"""
    name = sheet.title
    go_struct_name = to_pascal_case(name)

    lines = [f"package table\n", "const ("]
    for row in range(generate_common.BEGIN_ROW_IDX, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1)
        if cell.value is not None:
            const_name = f"{go_struct_name}TableID{int(cell.value)}"
            lines.append(f"    {const_name} = {int(cell.value)}")
    lines.append(")")
    return '\n'.join(lines)

def process_workbook(full_path: str, filename: str):
    """读取工作簿并生成 .go 文件"""
    try:
        workbook = openpyxl.load_workbook(full_path)
        sheetname = workbook.sheetnames[0]
        if sheetname not in generate_common.GEN_FILE_LIST:
            return

        sheet = workbook[sheetname]
        go_content = generate_go_const(sheet)
        output_file = os.path.join(GO_DIR, f"{sheetname.lower()}_table_id.go")
        generate_common.mywrite(go_content, output_file)
        logger.info(f"Generated Go const file: {output_file}")
    except Exception as e:
        logger.error(f"Failed to process {filename}: {e}")

def main():
    if not os.path.exists(GO_DIR):
        os.makedirs(GO_DIR)

    for filename in listdir(XLSX_DIR):
        full_path = os.path.join(XLSX_DIR, filename)
        if isfile(full_path) and filename.endswith('.xlsx'):
            process_workbook(full_path, filename)

if __name__ == "__main__":
    main()
