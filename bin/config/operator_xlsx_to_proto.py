#!/usr/bin/env python
# coding=utf-8

import os
import logging
import json
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook  # Use openpyxl for better support of .xlsx files

# Configure logging
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')

# Path configurations
excel_file_path = 'xlsx/operator/Operator.xlsx'
output_dir = 'generated/proto/operator'
temp_file_path = 'generated/proto/operator/temp_id_mapping.json'

# Ensure output directory exists
os.makedirs(output_dir, exist_ok=True)

# Use a lock for thread-safe file access
file_lock = threading.Lock()


def read_temp_id_mapping():
    """Reads existing ID mappings from temp file."""
    try:
        if os.path.exists(temp_file_path):
            with open(temp_file_path, 'r') as f:
                return json.load(f)
        else:
            return {}
    except Exception as e:
        logging.error(f"Error reading ID mapping file: {str(e)}")
        return {}


def write_temp_id_mapping(mapping):
    """Writes ID mappings to temp file."""
    try:
        with file_lock:
            with open(temp_file_path, 'w', encoding='utf-8') as f:
                json.dump(mapping, f, indent=2)
    except Exception as e:
        logging.error(f"Error writing ID mapping file: {str(e)}")


def read_excel_data(file_path):
    """Reads data from the first sheet of the provided Excel file."""
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active  # Only read the first sheet
        num_rows = sheet.max_row

        groups = {}
        current_group = None
        current_group_data = []
        global_row_id = 1

        for row_idx in range(18, num_rows + 1):  # Adjust for zero-based index and header rows
            row_cells = sheet[row_idx]

            if row_cells[0].value and row_cells[0].value.startswith('//'):
                group_name = row_cells[0].value.strip('/').strip()

                if current_group:
                    groups[current_group] = current_group_data

                current_group = group_name
                current_group_data = []
            else:
                if current_group:
                    enum_name = row_cells[0].value
                    if enum_name:
                        current_group_data.append((enum_name.strip(), global_row_id))
                        global_row_id += 1

        if current_group:
            groups[current_group] = current_group_data

        return groups

    except Exception as e:
        logging.error(f"Error reading Excel file: {str(e)}")
        return {}


def generate_proto_file(group_name, group_data, existing_id_mapping):
    """Generates Proto file for a given group."""
    try:
        proto_content = f"// Proto file for {group_name}\n"
        proto_content += 'syntax = "proto3";\n\n'
        proto_content += 'option go_package = "pb/game";\n\n'
        proto_content += f"enum {group_name} {{\n"
        proto_content += f"  k{group_name.capitalize()}OK = 0;\n"

        if group_name == "scene":
            proto_content += '  option allow_alias = true;\n\n'

        for enum_name, _ in group_data:
            if enum_name in existing_id_mapping:
                enum_id = existing_id_mapping[enum_name]
            else:
                enum_id = max(existing_id_mapping.values(), default=-1) + 1
                existing_id_mapping[enum_name] = enum_id

            enum_name_with_k = f"k{enum_name.strip()}"
            proto_content += f"  {enum_name_with_k} = {enum_id};\n"

        proto_content += '};\n'

        proto_file_path = os.path.join(output_dir, f"{group_name.lower()}_operator.proto")

        with file_lock:
            with open(proto_file_path, 'w', encoding='utf-8') as proto_file:
                proto_file.write(proto_content)

        logging.info(f"Proto enums file generated: {proto_file_path}")

        # Write updated ID mapping to temp file
        write_temp_id_mapping(existing_id_mapping)

    except Exception as e:
        logging.error(f"Error generating Proto file for group {group_name}: {str(e)}")


def generate_proto_files(groups):
    """Generates Proto files for all groups using ThreadPoolExecutor."""
    existing_id_mapping = read_temp_id_mapping()

    with ThreadPoolExecutor() as executor:
        futures = [executor.submit(generate_proto_file, group_name, group_data, existing_id_mapping)
                   for group_name, group_data in groups.items()]

        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                logging.error(f"Error occurred: {str(e)}")


def main():
    """Main function to read Excel data and generate Proto files."""
    groups = read_excel_data(excel_file_path)
    if groups:
        generate_proto_files(groups)
        logging.info("Proto generation completed.")


if __name__ == "__main__":
    main()
