import openpyxl

files = [
    r"data\Reward.xlsx",
    r"data\Mission.xlsx",
    r"data\Item.xlsx",
    r"data\Test.xlsx",
    r"data\Skill.xlsx",
    r"data\GlobalVariable.xlsx",
    r"data\TestMultiKey.xlsx",
]

for fpath in files:
    print()
    print("=" * 100)
    print("FILE:", fpath)
    print("=" * 100)
    wb = openpyxl.load_workbook(fpath)
    ws = wb[wb.sheetnames[0]]
    print("Sheet:", wb.sheetnames[0], " Rows:", ws.max_row, " Cols:", ws.max_column)
    for r in range(1, min(22, (ws.max_row or 1) + 1)):
        cells = []
        for c in range(1, min((ws.max_column or 0) + 1, 25)):
            v = ws.cell(row=r, column=c).value
            s = str(v)[:25] if v is not None else ""
            cells.append(s)
        print("  R%02d: %s" % (r, " | ".join(cells)))
